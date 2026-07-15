#!/usr/bin/env python3
"""
Multi-threaded Oracle vs PostgreSQL (Aurora) migration verification tool.

Comprehensive comparison per schema pair:
  * Row counts (all tables, union of both sides)
  * Indexes (matched by column signature)
  * Column data types (validated against AWS SCT default mapping rules)
  * Views, Materialized Views
  * Procedures & Functions (with cross-kind matching -- Oracle PROCEDURE
    frequently migrates to a Postgres FUNCTION)
  * Triggers, Sequences

Outputs:
  * db_comparison_report.xlsx  -- multi-sheet report
  * missing_indexes_postgres.sql  -- CREATE INDEX statements

Requirements:
    pip install oracledb psycopg2-binary openpyxl
"""

import concurrent.futures as cf
import logging
import re
import sys
import threading
import time
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

import oracledb
import psycopg2
from psycopg2 import pool as pg_pool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIG
# ============================================================================
ORACLE_CONFIG = {
    "user":     "oracle_user",
    "password": "oracle_password",
    "dsn":      "oracle-host.example.com:1521/ORCLPDB1",
}

POSTGRES_CONFIG = {
    "user":     "pg_user",
    "password": "pg_password",
    "host":     "aurora-cluster.cluster-xxxx.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "mydb",
}

SCHEMA_PAIRS = [
    ("SALES",   "sales"),
    ("HR",      "hr"),
    ("FINANCE", "finance"),
]

MAX_WORKERS_PER_DB = 16
POOL_SIZE          = 24
OUTPUT_FILE        = "db_comparison_report.xlsx"

# ---- Feature toggles -------------------------------------------------------
COMPARE_INDEXES         = True
COMPARE_COLUMNS         = True
COMPARE_VIEWS           = True
COMPARE_MVIEWS          = True
COMPARE_PROCS_AND_FUNCS = True
COMPARE_TRIGGERS        = True
COMPARE_SEQUENCES       = True

# ---- Missing-index DDL generation ------------------------------------------
GENERATE_MISSING_INDEX_DDL = True
MISSING_INDEX_DDL_FILE     = "missing_indexes_postgres.sql"
INDEX_DDL_CONCURRENTLY     = False

# ============================================================================
# Logging
# ============================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(threadName)-12s] %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("compare")


# ============================================================================
# Data models
# ============================================================================
@dataclass
class TableResult:
    oracle_schema: str
    pg_schema: str
    table_name: str
    in_oracle: bool = False
    in_postgres: bool = False
    ora_table_actual: str = ""
    pg_table_actual: str = ""
    oracle_count: Optional[int] = None
    pg_count: Optional[int] = None
    oracle_error: str = ""
    pg_error: str = ""

    @property
    def exists_both(self) -> str:
        return "YES" if (self.in_oracle and self.in_postgres) else "NO"

    @property
    def match(self) -> str:
        if not (self.in_oracle and self.in_postgres):
            return "NO"
        if self.oracle_count is None or self.pg_count is None:
            return "ERROR"
        return "YES" if self.oracle_count == self.pg_count else "NO"


@dataclass
class IndexInfo:
    schema: str
    table_key: str
    index_name: str
    unique: bool
    columns: Tuple[str, ...]
    descend: Tuple[bool, ...] = ()

    @property
    def signature(self) -> Tuple[str, ...]:
        return self.columns


@dataclass
class IndexRow:
    oracle_schema: str
    pg_schema: str
    table_name: str
    columns: str
    oracle_index: str
    pg_index: str
    oracle_unique: str
    pg_unique: str
    match: str
    remarks: str


@dataclass
class ColumnInfo:
    """One column, either Oracle or Postgres side."""
    schema: str
    table_key: str              # lower-cased table name
    column_name: str
    column_key: str             # lower-cased column name
    ordinal: int
    data_type_raw: str          # Oracle: 'NUMBER', 'VARCHAR2', ... ; PG: format_type() output
    length: Optional[int] = None
    precision: Optional[int] = None
    scale: Optional[int] = None
    nullable: bool = True


@dataclass
class ColumnRow:
    oracle_schema: str
    pg_schema: str
    table_name: str
    column_name: str
    oracle_type: str
    expected_pg_type: str
    actual_pg_type: str
    deviation: str        # SCT_MATCH / DEVIATION / MISSING_IN_PG / MISSING_IN_ORACLE
    match: str            # YES / NO
    remarks: str


@dataclass
class SimpleObjectRow:
    """Views, materialized views, sequences."""
    oracle_schema: str
    pg_schema: str
    object_name: str
    in_oracle: bool
    in_postgres: bool
    match: str
    remarks: str


@dataclass
class RoutineRow:
    """Procedures + functions (with cross-kind matching)."""
    oracle_schema: str
    pg_schema: str
    routine_name: str
    ora_kind: str     # PROCEDURE / FUNCTION / mixed / '-'
    pg_kind: str
    match: str        # YES / NO / CROSS-KIND
    remarks: str


@dataclass
class TriggerRow:
    oracle_schema: str
    pg_schema: str
    trigger_name: str
    ora_table: str
    pg_table: str
    in_oracle: bool
    in_postgres: bool
    match: str
    remarks: str


# ============================================================================
# Connection pools with semaphore-bounded waiting
# ============================================================================
class OraclePool:
    def __init__(self, cfg, size):
        self._sem = threading.BoundedSemaphore(size)
        self.pool = oracledb.create_pool(
            user=cfg["user"], password=cfg["password"], dsn=cfg["dsn"],
            min=2, max=size, increment=1,
            getmode=oracledb.POOL_GETMODE_WAIT,
        )

    def _acquire(self):
        self._sem.acquire()
        try:
            return self.pool.acquire()
        except Exception:
            self._sem.release()
            raise

    def _release(self, conn):
        try: self.pool.release(conn)
        finally: self._sem.release()

    def query_one(self, sql, params=None):
        conn = self._acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchone()
        finally: self._release(conn)

    def query_all(self, sql, params=None):
        conn = self._acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchall()
        finally: self._release(conn)

    def close(self): self.pool.close(force=True)


class PostgresPool:
    def __init__(self, cfg, size):
        self._sem = threading.BoundedSemaphore(size)
        self.pool = pg_pool.ThreadedConnectionPool(minconn=2, maxconn=size, **cfg)

    def _acquire(self):
        self._sem.acquire()
        try:
            return self.pool.getconn()
        except Exception:
            self._sem.release()
            raise

    def _release(self, conn):
        try:
            try: conn.rollback()
            finally: self.pool.putconn(conn)
        finally: self._sem.release()

    def _run(self, sql, params, fetch):
        conn = self._acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                return fetch(cur)
        finally: self._release(conn)

    def query_one(self, sql, params=None):
        return self._run(sql, params, lambda c: c.fetchone())

    def query_all(self, sql, params=None):
        return self._run(sql, params, lambda c: c.fetchall())

    def close(self): self.pool.closeall()


# ============================================================================
# AWS SCT type-mapping rules
# ============================================================================
_TS_PREC_RE = re.compile(r'TIMESTAMP\s*\((\d+)\)', re.IGNORECASE)


def expected_pg_type_for_oracle(data_type: str,
                                precision: Optional[int],
                                scale: Optional[int],
                                length: Optional[int]) -> str:
    """Return canonical (lower-case) expected Postgres type per AWS SCT defaults.

    Reference table:
      NUMBER(p,0), p ≤ 4       -> smallint
      NUMBER(p,0), p ≤ 9       -> integer
      NUMBER(p,0), p ≤ 18      -> bigint
      NUMBER(p,0), p > 18      -> numeric(p,0)
      NUMBER(p,s), s > 0       -> numeric(p,s)
      NUMBER (no precision)    -> double precision
      FLOAT(p)                 -> real if p <= 24 else double precision
      BINARY_FLOAT             -> real
      BINARY_DOUBLE            -> double precision
      VARCHAR2(n)/NVARCHAR2(n) -> varchar(n)
      CHAR(n)/NCHAR(n)         -> char(n)
      CLOB/NCLOB/LONG          -> text
      BLOB/RAW/LONG RAW        -> bytea
      ROWID/UROWID             -> varchar(255)
      XMLTYPE                  -> xml
      DATE                     -> timestamp(0)
      TIMESTAMP(n)             -> timestamp(n)     (bare 'timestamp' if n=6)
      TIMESTAMP(n) WITH [LOCAL] TIME ZONE -> timestamp(n) with time zone
      INTERVAL ...             -> interval
    """
    t = (data_type or "").upper().strip()
    p = int(precision) if precision is not None else None
    s = int(scale) if scale is not None else None
    L = int(length) if length is not None else None

    if t == 'NUMBER':
        if p is None:
            return 'double precision'
        if s and s > 0:
            return f'numeric({p},{s})'
        if p <= 4:  return 'smallint'
        if p <= 9:  return 'integer'
        if p <= 18: return 'bigint'
        return f'numeric({p},0)'

    if t == 'FLOAT':
        # Oracle FLOAT(p) uses binary precision. SCT: p <= 24 -> real else double.
        return 'real' if (p or 126) <= 24 else 'double precision'

    if t == 'BINARY_FLOAT':  return 'real'
    if t == 'BINARY_DOUBLE': return 'double precision'

    if t in ('VARCHAR2', 'NVARCHAR2'):
        return f'varchar({L})' if L else 'varchar'
    if t in ('CHAR', 'NCHAR'):
        return f'char({L})' if L else 'char'

    if t in ('CLOB', 'NCLOB', 'LONG'):   return 'text'
    if t in ('BLOB', 'RAW', 'LONG RAW'): return 'bytea'
    if t in ('ROWID', 'UROWID'):          return 'varchar(255)'
    if t == 'XMLTYPE':                    return 'xml'
    if t == 'DATE':                       return 'timestamp(0)'

    if 'TIMESTAMP' in t:
        # Precision may be in the type string (TIMESTAMP(6)) or in data_scale.
        m = _TS_PREC_RE.search(t)
        prec = int(m.group(1)) if m else (s if s is not None else 6)
        tz = ' with time zone' if ('WITH TIME ZONE' in t or 'LOCAL TIME ZONE' in t) else ''
        return f'timestamp{tz}' if prec == 6 else f'timestamp({prec}){tz}'

    if 'INTERVAL' in t:
        return 'interval'

    # Unknown -- return normalized raw
    return t.lower()


def normalize_pg_type(t: str) -> str:
    """Canonicalize a Postgres format_type() output for comparison."""
    t = (t or "").lower().strip()
    t = t.replace("character varying", "varchar")
    if t.startswith("character(") or t == "character":
        t = "char" + t[len("character"):]
    t = t.replace(" without time zone", "")
    while "  " in t:
        t = t.replace("  ", " ")
    return t.strip()


def format_oracle_type_display(ci: ColumnInfo) -> str:
    """Human-readable Oracle type for the report."""
    t = ci.data_type_raw.upper()
    if t == 'NUMBER':
        if ci.precision is None:
            return 'NUMBER'
        if ci.scale and ci.scale > 0:
            return f'NUMBER({ci.precision},{ci.scale})'
        return f'NUMBER({ci.precision})'
    if t in ('VARCHAR2', 'NVARCHAR2', 'CHAR', 'NCHAR', 'RAW'):
        return f'{t}({ci.length})' if ci.length else t
    if t == 'FLOAT' and ci.precision is not None:
        return f'FLOAT({ci.precision})'
    return t


# ============================================================================
# Table + column metadata
# ============================================================================
def get_oracle_tables(orapool, schema):
    rows = orapool.query_all(
        """
        SELECT table_name
        FROM all_tables
        WHERE owner = :owner
          AND table_name NOT LIKE 'BIN$%'
          AND (nested = 'NO' OR nested IS NULL)
          AND iot_type IS NULL
        """,
        {"owner": schema.upper()},
    )
    tables = {r[0] for r in rows}
    log.info("Oracle schema %s: %d tables found", schema, len(tables))
    return tables


def get_postgres_tables(pgpool, schema):
    # Exclude partition / inheritance children -- Oracle keeps partitions out
    # of all_tables so we drop the Postgres child rows for parity.
    rows = pgpool.query_all(
        """
        SELECT c.relname
        FROM pg_class c
        JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND c.relkind IN ('r', 'p', 'f')
          AND c.relname NOT LIKE 'pg_%%'
          AND NOT EXISTS (SELECT 1 FROM pg_inherits WHERE inhrelid = c.oid)
        """,
        (schema,),
    )
    tables = {r[0] for r in rows}
    log.info("Postgres schema %s: %d tables found", schema, len(tables))
    return tables


def get_oracle_columns(orapool, schema):
    rows = orapool.query_all(
        """
        SELECT table_name, column_name, column_id, data_type,
               data_length, data_precision, data_scale, nullable
        FROM all_tab_columns
        WHERE owner = :owner
          AND table_name NOT LIKE 'BIN$%'
        ORDER BY table_name, column_id
        """,
        {"owner": schema.upper()},
    )
    result = []
    for tbl, col, cid, dtype, dlen, dprec, dscale, nullable in rows:
        result.append(ColumnInfo(
            schema=schema, table_key=tbl.lower(),
            column_name=col, column_key=col.lower(),
            ordinal=int(cid) if cid is not None else 0,
            data_type_raw=dtype,
            length=int(dlen) if dlen is not None else None,
            precision=int(dprec) if dprec is not None else None,
            scale=int(dscale) if dscale is not None else None,
            nullable=(str(nullable).upper() == 'Y'),
        ))
    log.info("Oracle schema %s: %d columns", schema, len(result))
    return result


def get_postgres_columns(pgpool, schema):
    rows = pgpool.query_all(
        """
        SELECT c.relname                          AS table_name,
               a.attname                          AS column_name,
               a.attnum                           AS ordinal,
               format_type(a.atttypid, a.atttypmod) AS data_type,
               NOT a.attnotnull                   AS nullable
        FROM pg_class c
        JOIN pg_namespace n ON n.oid = c.relnamespace
        JOIN pg_attribute a ON a.attrelid = c.oid
        WHERE n.nspname = %s
          AND c.relkind IN ('r', 'p', 'f')
          AND a.attnum > 0
          AND NOT a.attisdropped
          AND NOT EXISTS (SELECT 1 FROM pg_inherits WHERE inhrelid = c.oid)
        ORDER BY c.relname, a.attnum
        """,
        (schema,),
    )
    result = []
    for tbl, col, ordinal, dtype, nullable in rows:
        result.append(ColumnInfo(
            schema=schema, table_key=tbl.lower(),
            column_name=col, column_key=col.lower(),
            ordinal=int(ordinal),
            data_type_raw=dtype or "",
            nullable=bool(nullable),
        ))
    log.info("Postgres schema %s: %d columns", schema, len(result))
    return result


# ============================================================================
# Index metadata
# ============================================================================
def get_oracle_indexes(orapool, schema):
    rows = orapool.query_all(
        """
        SELECT i.table_name, i.index_name, i.uniqueness,
               c.column_name, c.column_position, c.descend
        FROM all_indexes i
        JOIN all_ind_columns c
          ON i.owner = c.index_owner AND i.index_name = c.index_name
        WHERE i.owner = :owner
          AND i.index_name NOT LIKE 'BIN$%'
          AND i.dropped = 'NO'
          AND i.index_type NOT LIKE 'IOT%'
          AND i.index_type NOT LIKE 'LOB%'
          AND i.table_name NOT LIKE 'BIN$%'
        ORDER BY i.table_name, i.index_name, c.column_position
        """,
        {"owner": schema.upper()},
    )
    bucket = {}
    for tbl, idx, uniq, col, _pos, descend in rows:
        key = (tbl, idx)
        info = bucket.get(key)
        if info is None:
            info = IndexInfo(schema=schema, table_key=tbl.lower(),
                             index_name=idx, unique=(uniq == "UNIQUE"),
                             columns=[], descend=[])
            bucket[key] = info
        info.columns.append((col or "").lower())
        info.descend.append(str(descend or "").upper() == "DESC")
    result = []
    for info in bucket.values():
        info.columns = tuple(info.columns)
        info.descend = tuple(info.descend)
        result.append(info)
    log.info("Oracle schema %s: %d indexes", schema, len(result))
    return result


def get_postgres_indexes(pgpool, schema):
    rows = pgpool.query_all(
        """
        SELECT tc.relname, ic.relname, ix.indisunique,
               col.attname, col.ord
        FROM pg_index ix
        JOIN pg_class ic ON ic.oid = ix.indexrelid
        JOIN pg_class tc ON tc.oid = ix.indrelid
        JOIN pg_namespace n ON n.oid = tc.relnamespace
        CROSS JOIN LATERAL (
            SELECT COALESCE(a.attname, '<expr>') AS attname, k.ord
            FROM unnest(ix.indkey) WITH ORDINALITY AS k(attnum, ord)
            LEFT JOIN pg_attribute a
                   ON a.attrelid = tc.oid AND a.attnum = k.attnum
            WHERE k.ord <= ix.indnkeyatts
        ) col
        WHERE n.nspname = %s
          AND tc.relkind IN ('r', 'p')
          AND tc.relname NOT LIKE 'pg_%%'
          AND NOT EXISTS (SELECT 1 FROM pg_inherits WHERE inhrelid = tc.oid)
        ORDER BY tc.relname, ic.relname, col.ord
        """,
        (schema,),
    )
    bucket = {}
    for tbl, idx, uniq, col, _pos in rows:
        key = (tbl, idx)
        info = bucket.get(key)
        if info is None:
            info = IndexInfo(schema=schema, table_key=tbl.lower(),
                             index_name=idx, unique=bool(uniq),
                             columns=[])
            bucket[key] = info
        info.columns.append((col or "").lower())
    result = []
    for info in bucket.values():
        info.columns = tuple(info.columns)
        result.append(info)
    log.info("Postgres schema %s: %d indexes", schema, len(result))
    return result


# ============================================================================
# Views / MViews / Sequences metadata
# ============================================================================
def get_oracle_names(orapool, sql, schema):
    return {r[0] for r in orapool.query_all(sql, {"owner": schema.upper()})}


def get_postgres_names(pgpool, sql, schema):
    return {r[0] for r in pgpool.query_all(sql, (schema,))}


def get_oracle_views(orapool, schema):
    v = get_oracle_names(
        orapool, "SELECT view_name FROM all_views WHERE owner = :owner", schema)
    log.info("Oracle schema %s: %d views", schema, len(v))
    return v


def get_postgres_views(pgpool, schema):
    v = get_postgres_names(
        pgpool,
        """SELECT c.relname FROM pg_class c
           JOIN pg_namespace n ON n.oid = c.relnamespace
           WHERE n.nspname = %s AND c.relkind = 'v'""",
        schema)
    log.info("Postgres schema %s: %d views", schema, len(v))
    return v


def get_oracle_mviews(orapool, schema):
    v = get_oracle_names(
        orapool, "SELECT mview_name FROM all_mviews WHERE owner = :owner", schema)
    log.info("Oracle schema %s: %d mviews", schema, len(v))
    return v


def get_postgres_mviews(pgpool, schema):
    v = get_postgres_names(
        pgpool,
        """SELECT c.relname FROM pg_class c
           JOIN pg_namespace n ON n.oid = c.relnamespace
           WHERE n.nspname = %s AND c.relkind = 'm'""",
        schema)
    log.info("Postgres schema %s: %d mviews", schema, len(v))
    return v


def get_oracle_sequences(orapool, schema):
    v = {r[0] for r in orapool.query_all(
        "SELECT sequence_name FROM all_sequences WHERE sequence_owner = :owner",
        {"owner": schema.upper()})}
    log.info("Oracle schema %s: %d sequences", schema, len(v))
    return v


def get_postgres_sequences(pgpool, schema):
    """Return dict {name: is_owned_by_column} so we can flag implicit SERIAL sequences."""
    rows = pgpool.query_all(
        """
        SELECT c.relname,
               EXISTS(SELECT 1 FROM pg_depend d
                      WHERE d.objid = c.oid AND d.deptype = 'a') AS owned
        FROM pg_class c
        JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s AND c.relkind = 'S'
        """,
        (schema,),
    )
    result = {r[0]: bool(r[1]) for r in rows}
    log.info("Postgres schema %s: %d sequences", schema, len(result))
    return result


# ============================================================================
# Procedures + Functions metadata
# ============================================================================
def get_oracle_routines(orapool, schema):
    """Return list of (name, kind) where kind in {'PROCEDURE', 'FUNCTION'}."""
    rows = orapool.query_all(
        """
        SELECT object_name, object_type
        FROM all_objects
        WHERE owner = :owner
          AND object_type IN ('PROCEDURE', 'FUNCTION')
          AND status = 'VALID'
        """,
        {"owner": schema.upper()},
    )
    result = [(r[0], r[1]) for r in rows]
    log.info("Oracle schema %s: %d procedures/functions", schema, len(result))
    return result


def get_postgres_routines(pgpool, schema):
    rows = pgpool.query_all(
        """
        SELECT DISTINCT p.proname,
               CASE p.prokind
                    WHEN 'p' THEN 'PROCEDURE'
                    WHEN 'f' THEN 'FUNCTION'
                    WHEN 'a' THEN 'AGGREGATE'
                    WHEN 'w' THEN 'WINDOW'
                    ELSE 'OTHER'
               END AS kind
        FROM pg_proc p
        JOIN pg_namespace n ON n.oid = p.pronamespace
        WHERE n.nspname = %s
          AND p.prokind IN ('p', 'f')
        """,
        (schema,),
    )
    result = [(r[0], r[1]) for r in rows]
    log.info("Postgres schema %s: %d procedures/functions", schema, len(result))
    return result


# ============================================================================
# Triggers metadata
# ============================================================================
def get_oracle_triggers(orapool, schema):
    """Return list of (trigger_name, table_name)."""
    rows = orapool.query_all(
        """
        SELECT trigger_name, table_name
        FROM all_triggers
        WHERE owner = :owner
          AND status = 'ENABLED'
        """,
        {"owner": schema.upper()},
    )
    result = [(r[0], r[1] or "") for r in rows]
    log.info("Oracle schema %s: %d triggers", schema, len(result))
    return result


def get_postgres_triggers(pgpool, schema):
    rows = pgpool.query_all(
        """
        SELECT t.tgname, c.relname
        FROM pg_trigger t
        JOIN pg_class c ON c.oid = t.tgrelid
        JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND NOT t.tgisinternal
        """,
        (schema,),
    )
    result = [(r[0], r[1] or "") for r in rows]
    log.info("Postgres schema %s: %d triggers", schema, len(result))
    return result


# ============================================================================
# Row-count workers
# ============================================================================
PG_UNDEFINED_TABLE = "42P01"
ORA_TABLE_OR_VIEW_NOT_EXIST = 942


def count_oracle(orapool, result: TableResult):
    sql = f'SELECT COUNT(*) FROM "{result.oracle_schema.upper()}"."{result.ora_table_actual}"'
    t0 = time.time()
    try:
        row = orapool.query_one(sql)
        result.oracle_count = int(row[0])
        log.info("Oracle  %s.%s = %s (%.1fs)",
                 result.oracle_schema, result.ora_table_actual,
                 f"{result.oracle_count:,}", time.time() - t0)
    except oracledb.DatabaseError as e:
        err, = e.args
        if getattr(err, "code", None) == ORA_TABLE_OR_VIEW_NOT_EXIST:
            result.in_oracle = False
            result.oracle_error = "Table not found at count time"
        else:
            result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle  %s.%s FAILED: %s",
                  result.oracle_schema, result.ora_table_actual, result.oracle_error)
    except Exception as e:
        result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle  %s.%s FAILED: %s",
                  result.oracle_schema, result.ora_table_actual, result.oracle_error)


def count_postgres(pgpool, result: TableResult):
    sql = f'SELECT COUNT(*) FROM "{result.pg_schema}"."{result.pg_table_actual}"'
    t0 = time.time()
    try:
        row = pgpool.query_one(sql)
        result.pg_count = int(row[0])
        log.info("Postgres %s.%s = %s (%.1fs)",
                 result.pg_schema, result.pg_table_actual,
                 f"{result.pg_count:,}", time.time() - t0)
    except psycopg2.Error as e:
        if getattr(e, "pgcode", None) == PG_UNDEFINED_TABLE:
            result.in_postgres = False
            result.pg_error = "Table not found at count time"
        else:
            result.pg_error = str(e).splitlines()[0][:200]
        log.error("Postgres %s.%s FAILED: %s",
                  result.pg_schema, result.pg_table_actual, result.pg_error)
    except Exception as e:
        result.pg_error = str(e).splitlines()[0][:200]
        log.error("Postgres %s.%s FAILED: %s",
                  result.pg_schema, result.pg_table_actual, result.pg_error)


# ============================================================================
# Index comparison
# ============================================================================
def compare_indexes_for_pair(ora_schema, pg_schema, common_tables,
                             ora_indexes, pg_indexes):
    ora_by_table = defaultdict(list)
    for idx in ora_indexes:
        if idx.table_key in common_tables:
            ora_by_table[idx.table_key].append(idx)
    pg_by_table = defaultdict(list)
    for idx in pg_indexes:
        if idx.table_key in common_tables:
            pg_by_table[idx.table_key].append(idx)

    rows = []
    for table_key in sorted(common_tables):
        ora_list = ora_by_table.get(table_key, [])
        pg_list = pg_by_table.get(table_key, [])
        if not ora_list and not pg_list:
            continue
        ora_by_sig = defaultdict(list)
        for i in ora_list: ora_by_sig[i.signature].append(i)
        pg_by_sig = defaultdict(list)
        for i in pg_list: pg_by_sig[i.signature].append(i)
        display_table = ora_list[0].table_key if ora_list else pg_list[0].table_key
        for sig in sorted(set(ora_by_sig) | set(pg_by_sig)):
            o = ora_by_sig.get(sig, []); p = pg_by_sig.get(sig, [])
            for k in range(max(len(o), len(p))):
                oi = o[k] if k < len(o) else None
                pi = p[k] if k < len(p) else None
                remarks = []
                if oi and not pi:
                    match = "NO"; remarks.append("Missing in Postgres")
                elif pi and not oi:
                    match = "NO"; remarks.append("Missing in Oracle")
                else:
                    match = "YES" if oi.unique == pi.unique else "NO"
                    if oi.unique != pi.unique: remarks.append("Uniqueness differs")
                if sig and any(c == "<expr>" or c.startswith("sys_nc") for c in sig):
                    remarks.append("Contains expression/function column")
                rows.append(IndexRow(
                    oracle_schema=ora_schema, pg_schema=pg_schema,
                    table_name=display_table,
                    columns=", ".join(sig) if sig else "(none)",
                    oracle_index=oi.index_name if oi else "-",
                    pg_index=pi.index_name if pi else "-",
                    oracle_unique=("YES" if oi.unique else "NO") if oi else "-",
                    pg_unique=("YES" if pi.unique else "NO") if pi else "-",
                    match=match, remarks="; ".join(remarks),
                ))
    return rows


# ============================================================================
# Column comparison
# ============================================================================
def compare_columns_for_pair(ora_schema, pg_schema, common_tables,
                             ora_cols, pg_cols, pg_actual_by_key):
    ora_by_table = defaultdict(list)
    for c in ora_cols:
        if c.table_key in common_tables:
            ora_by_table[c.table_key].append(c)
    pg_by_table = defaultdict(list)
    for c in pg_cols:
        if c.table_key in common_tables:
            pg_by_table[c.table_key].append(c)

    rows = []
    for tk in sorted(common_tables):
        pg_table = pg_actual_by_key.get(tk, tk)
        ora_map = {c.column_key: c for c in ora_by_table.get(tk, [])}
        pg_map = {c.column_key: c for c in pg_by_table.get(tk, [])}
        all_cols = sorted(set(ora_map) | set(pg_map),
                          key=lambda k: (ora_map.get(k, pg_map.get(k)).ordinal, k))
        for ck in all_cols:
            oc = ora_map.get(ck)
            pc = pg_map.get(ck)
            remarks = []

            if oc and not pc:
                expected = expected_pg_type_for_oracle(
                    oc.data_type_raw, oc.precision, oc.scale, oc.length)
                rows.append(ColumnRow(
                    oracle_schema=ora_schema, pg_schema=pg_schema,
                    table_name=pg_table, column_name=oc.column_name,
                    oracle_type=format_oracle_type_display(oc),
                    expected_pg_type=expected, actual_pg_type="-",
                    deviation="MISSING_IN_PG", match="NO",
                    remarks="Column missing in Postgres",
                ))
                continue

            if pc and not oc:
                actual = normalize_pg_type(pc.data_type_raw)
                rows.append(ColumnRow(
                    oracle_schema=ora_schema, pg_schema=pg_schema,
                    table_name=pg_table, column_name=pc.column_name,
                    oracle_type="-", expected_pg_type="-",
                    actual_pg_type=actual,
                    deviation="MISSING_IN_ORACLE", match="NO",
                    remarks="Column missing in Oracle",
                ))
                continue

            # Both sides present
            expected = expected_pg_type_for_oracle(
                oc.data_type_raw, oc.precision, oc.scale, oc.length)
            actual = normalize_pg_type(pc.data_type_raw)
            if expected == actual:
                deviation = "SCT_MATCH"; match = "YES"
            else:
                deviation = "DEVIATION"; match = "NO"
                remarks.append(f"Expected '{expected}', got '{actual}'")

            if oc.nullable != pc.nullable:
                remarks.append(
                    f"Nullability differs (Oracle={'Y' if oc.nullable else 'N'}, "
                    f"Postgres={'Y' if pc.nullable else 'N'})"
                )

            rows.append(ColumnRow(
                oracle_schema=ora_schema, pg_schema=pg_schema,
                table_name=pg_table, column_name=oc.column_name,
                oracle_type=format_oracle_type_display(oc),
                expected_pg_type=expected, actual_pg_type=actual,
                deviation=deviation, match=match,
                remarks="; ".join(remarks),
            ))
    return rows


# ============================================================================
# Simple object comparison (views, mviews)
# ============================================================================
def compare_simple_objects(ora_schema, pg_schema, ora_names, pg_names):
    ora_map = {n.lower(): n for n in ora_names}
    pg_map = {n.lower(): n for n in pg_names}
    rows = []
    for key in sorted(set(ora_map) | set(pg_map)):
        in_ora = key in ora_map
        in_pg = key in pg_map
        remarks = []
        if not in_ora: remarks.append("Missing in Oracle")
        if not in_pg:  remarks.append("Missing in Postgres")
        rows.append(SimpleObjectRow(
            oracle_schema=ora_schema, pg_schema=pg_schema,
            object_name=ora_map.get(key, pg_map.get(key, key)),
            in_oracle=in_ora, in_postgres=in_pg,
            match="YES" if (in_ora and in_pg) else "NO",
            remarks="; ".join(remarks),
        ))
    return rows


def compare_sequences(ora_schema, pg_schema, ora_names, pg_names_owned):
    """pg_names_owned: dict {name: is_owned_by_column}."""
    ora_map = {n.lower(): n for n in ora_names}
    pg_map = {n.lower(): n for n in pg_names_owned}
    rows = []
    for key in sorted(set(ora_map) | set(pg_map)):
        in_ora = key in ora_map
        in_pg = key in pg_map
        remarks = []
        if not in_ora: remarks.append("Missing in Oracle")
        if not in_pg:  remarks.append("Missing in Postgres")
        if in_pg and pg_names_owned.get(pg_map[key], False):
            remarks.append("Postgres sequence is owned by a column (implicit SERIAL/IDENTITY)")
        rows.append(SimpleObjectRow(
            oracle_schema=ora_schema, pg_schema=pg_schema,
            object_name=ora_map.get(key, pg_map.get(key, key)),
            in_oracle=in_ora, in_postgres=in_pg,
            match="YES" if (in_ora and in_pg) else "NO",
            remarks="; ".join(remarks),
        ))
    return rows


# ============================================================================
# Routines (procedures + functions) with cross-kind matching
# ============================================================================
def compare_routines(ora_schema, pg_schema, ora_routines, pg_routines):
    ora_map = defaultdict(set)
    for name, kind in ora_routines: ora_map[name.lower()].add(kind)
    pg_map = defaultdict(set)
    for name, kind in pg_routines: pg_map[name.lower()].add(kind)

    rows = []
    for key in sorted(set(ora_map) | set(pg_map)):
        ora_kinds = ora_map.get(key, set())
        pg_kinds = pg_map.get(key, set())
        remarks = []
        if not ora_kinds:
            match = "NO"; remarks.append("Missing in Oracle")
        elif not pg_kinds:
            match = "NO"; remarks.append("Missing in Postgres")
        elif ora_kinds & pg_kinds:
            match = "YES"
        else:
            match = "CROSS-KIND"
            remarks.append(
                f"Kind changed: Oracle {'/'.join(sorted(ora_kinds))} "
                f"vs Postgres {'/'.join(sorted(pg_kinds))}"
            )
        # Use a nice display name (prefer Oracle case)
        display_name = key
        for name, _ in ora_routines:
            if name.lower() == key:
                display_name = name; break
        else:
            for name, _ in pg_routines:
                if name.lower() == key:
                    display_name = name; break
        rows.append(RoutineRow(
            oracle_schema=ora_schema, pg_schema=pg_schema,
            routine_name=display_name,
            ora_kind="/".join(sorted(ora_kinds)) if ora_kinds else "-",
            pg_kind="/".join(sorted(pg_kinds)) if pg_kinds else "-",
            match=match, remarks="; ".join(remarks),
        ))
    return rows


# ============================================================================
# Triggers
# ============================================================================
def compare_triggers(ora_schema, pg_schema, ora_triggers, pg_triggers):
    ora_map = {name.lower(): (name, tbl) for name, tbl in ora_triggers}
    pg_map = {name.lower(): (name, tbl) for name, tbl in pg_triggers}
    rows = []
    for key in sorted(set(ora_map) | set(pg_map)):
        in_ora = key in ora_map
        in_pg = key in pg_map
        remarks = []
        if not in_ora: remarks.append("Missing in Oracle")
        if not in_pg:  remarks.append("Missing in Postgres")
        # Table mismatch check
        if in_ora and in_pg:
            ora_tbl = ora_map[key][1]
            pg_tbl = pg_map[key][1]
            if ora_tbl.lower() != pg_tbl.lower():
                remarks.append(f"On different tables (Oracle={ora_tbl}, Postgres={pg_tbl})")
        display_name = ora_map.get(key, pg_map.get(key, (key, "")))[0]
        rows.append(TriggerRow(
            oracle_schema=ora_schema, pg_schema=pg_schema,
            trigger_name=display_name,
            ora_table=ora_map.get(key, ("", ""))[1] if in_ora else "-",
            pg_table=pg_map.get(key, ("", ""))[1] if in_pg else "-",
            in_oracle=in_ora, in_postgres=in_pg,
            match="YES" if (in_ora and in_pg
                            and (not remarks or all("different tables" not in r for r in remarks))) else "NO",
            remarks="; ".join(remarks),
        ))
    return rows


# ============================================================================
# Missing-index DDL
# ============================================================================
def _safe_pg_index_name(base_name, taken):
    name = base_name.lower()[:63]
    if name not in taken: return name
    for i in range(1, 1000):
        suffix = f"_{i}"
        candidate = name[:63 - len(suffix)] + suffix
        if candidate not in taken: return candidate
    return name


def _build_create_index_stmt(pg_schema, pg_table, oi, taken_names):
    if any(c == "<expr>" or c.startswith("sys_nc") or not c for c in oi.columns):
        return (f"-- SKIP: {oi.index_name} on {pg_schema}.{pg_table} is a "
                f"function-based / expression index; translate manually.",
                None, "function-based / expression index")
    if oi.index_name.upper().startswith("SYS_C") and oi.unique:
        return (f"-- SKIP: {oi.index_name} on {pg_schema}.{pg_table} looks like a "
                f"system-generated PK/UNIQUE constraint index; add the constraint "
                f"via ALTER TABLE instead.",
                None, "system-generated constraint index")
    new_name = _safe_pg_index_name(oi.index_name, taken_names)
    unique = "UNIQUE " if oi.unique else ""
    concur = "CONCURRENTLY " if INDEX_DDL_CONCURRENTLY else ""
    parts = []
    for i, col in enumerate(oi.columns):
        d = " DESC" if i < len(oi.descend) and oi.descend[i] else ""
        parts.append(f'"{col}"{d}')
    stmt = (f'CREATE {unique}INDEX {concur}"{new_name}" '
            f'ON "{pg_schema}"."{pg_table}" ({", ".join(parts)});')
    return stmt, new_name, None


def generate_missing_pg_index_ddl(bundles, path):
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    generated = 0; skipped = []; body = []
    existing = defaultdict(set)
    for b in bundles:
        for pi in b.pg_indexes:
            existing[b.pg_schema].add(pi.index_name.lower())
    invented = defaultdict(set)

    for b in bundles:
        common = ({t.lower() for t in b.ora_tables}
                  & {t.lower() for t in b.pg_tables})
        pg_actual = {t.lower(): t for t in b.pg_tables}
        pg_sigs = defaultdict(set)
        for pi in b.pg_indexes:
            if pi.table_key in common:
                pg_sigs[pi.table_key].add(pi.signature)
        missing_by_table = defaultdict(list)
        for oi in b.ora_indexes:
            if oi.table_key in common and oi.signature not in pg_sigs[oi.table_key]:
                missing_by_table[oi.table_key].append(oi)
        if not missing_by_table: continue
        body.append("")
        body.append("-- ------------------------------------------------------------")
        body.append(f"-- Oracle {b.ora_schema}  -->  Postgres {b.pg_schema}")
        body.append("-- ------------------------------------------------------------")
        for tk in sorted(missing_by_table):
            pg_table = pg_actual.get(tk, tk)
            body.append(""); body.append(f"-- Table: {b.pg_schema}.{pg_table}")
            for oi in missing_by_table[tk]:
                taken = existing[b.pg_schema] | invented[b.pg_schema]
                stmt, new_name, skip_reason = _build_create_index_stmt(
                    b.pg_schema, pg_table, oi, taken)
                body.append(stmt)
                if skip_reason:
                    skipped.append((b.pg_schema, pg_table, oi.index_name, skip_reason))
                else:
                    invented[b.pg_schema].add(new_name); generated += 1

    header = [
        "-- ============================================================",
        "-- Missing Postgres indexes -- CREATE INDEX statements",
        f"-- Generated: {now}",
        f"-- Statements generated: {generated}",
        f"-- Statements skipped:   {len(skipped)}",
        "--",
        "-- Match criterion: Oracle index whose ordered column signature",
        "-- is NOT present on any Postgres index of the same table.",
        "-- Review before running in production.",
        ("-- CONCURRENTLY mode is ENABLED (run outside a transaction; one at a time)."
         if INDEX_DDL_CONCURRENTLY else
         "-- For online use, set INDEX_DDL_CONCURRENTLY = True and re-run."),
        "-- ============================================================",
    ]
    footer = []
    if skipped:
        footer.append(""); footer.append("-- ============================================================")
        footer.append("-- Skipped indexes (require manual translation)")
        footer.append("-- ============================================================")
        for sch, tbl, name, reason in skipped:
            footer.append(f"--   {sch}.{tbl} :: {name}  --  {reason}")
    with open(path, "w") as f:
        f.write("\n".join(header + body + footer) + "\n")
    log.info("Missing-index DDL written: %s (generated: %d, skipped: %d)",
             path, generated, len(skipped))
    return generated, len(skipped)


# ============================================================================
# Orchestration
# ============================================================================
@dataclass
class SchemaBundle:
    ora_schema: str
    pg_schema: str
    ora_tables: set = field(default_factory=set)
    pg_tables: set = field(default_factory=set)
    ora_indexes: List[IndexInfo] = field(default_factory=list)
    pg_indexes: List[IndexInfo] = field(default_factory=list)
    ora_columns: List[ColumnInfo] = field(default_factory=list)
    pg_columns: List[ColumnInfo] = field(default_factory=list)
    ora_views: set = field(default_factory=set)
    pg_views: set = field(default_factory=set)
    ora_mviews: set = field(default_factory=set)
    pg_mviews: set = field(default_factory=set)
    ora_routines: list = field(default_factory=list)
    pg_routines: list = field(default_factory=list)
    ora_triggers: list = field(default_factory=list)
    pg_triggers: list = field(default_factory=list)
    ora_sequences: set = field(default_factory=set)
    pg_sequences: dict = field(default_factory=dict)  # name -> is_owned
    # Comparison results
    table_results: List[TableResult] = field(default_factory=list)
    index_results: List[IndexRow] = field(default_factory=list)
    column_results: List[ColumnRow] = field(default_factory=list)
    view_results: List[SimpleObjectRow] = field(default_factory=list)
    mview_results: List[SimpleObjectRow] = field(default_factory=list)
    routine_results: List[RoutineRow] = field(default_factory=list)
    trigger_results: List[TriggerRow] = field(default_factory=list)
    sequence_results: List[SimpleObjectRow] = field(default_factory=list)


def gather_metadata(orapool, pgpool, meta_ex):
    bundles = [SchemaBundle(o, p) for o, p in SCHEMA_PAIRS]
    futs = {}

    def submit(pool_fn, pool, schema, attr, target):
        futs[meta_ex.submit(pool_fn, pool, schema)] = (target, attr)

    for b in bundles:
        submit(get_oracle_tables,   orapool, b.ora_schema, "ora_tables", b)
        submit(get_postgres_tables, pgpool,  b.pg_schema,  "pg_tables",  b)
        if COMPARE_INDEXES:
            submit(get_oracle_indexes,   orapool, b.ora_schema, "ora_indexes", b)
            submit(get_postgres_indexes, pgpool,  b.pg_schema,  "pg_indexes",  b)
        if COMPARE_COLUMNS:
            submit(get_oracle_columns,   orapool, b.ora_schema, "ora_columns", b)
            submit(get_postgres_columns, pgpool,  b.pg_schema,  "pg_columns",  b)
        if COMPARE_VIEWS:
            submit(get_oracle_views,   orapool, b.ora_schema, "ora_views", b)
            submit(get_postgres_views, pgpool,  b.pg_schema,  "pg_views",  b)
        if COMPARE_MVIEWS:
            submit(get_oracle_mviews,   orapool, b.ora_schema, "ora_mviews", b)
            submit(get_postgres_mviews, pgpool,  b.pg_schema,  "pg_mviews",  b)
        if COMPARE_PROCS_AND_FUNCS:
            submit(get_oracle_routines,   orapool, b.ora_schema, "ora_routines", b)
            submit(get_postgres_routines, pgpool,  b.pg_schema,  "pg_routines",  b)
        if COMPARE_TRIGGERS:
            submit(get_oracle_triggers,   orapool, b.ora_schema, "ora_triggers", b)
            submit(get_postgres_triggers, pgpool,  b.pg_schema,  "pg_triggers",  b)
        if COMPARE_SEQUENCES:
            submit(get_oracle_sequences,   orapool, b.ora_schema, "ora_sequences", b)
            submit(get_postgres_sequences, pgpool,  b.pg_schema,  "pg_sequences",  b)

    for fut in cf.as_completed(futs):
        target, attr = futs[fut]
        setattr(target, attr, fut.result())

    for b in bundles:
        ora_map = {t.lower(): t for t in b.ora_tables}
        pg_map = {t.lower(): t for t in b.pg_tables}
        for key in sorted(set(ora_map) | set(pg_map)):
            b.table_results.append(TableResult(
                oracle_schema=b.ora_schema, pg_schema=b.pg_schema,
                table_name=ora_map.get(key, pg_map.get(key, key)),
                in_oracle=key in ora_map, in_postgres=key in pg_map,
                ora_table_actual=ora_map.get(key, ""),
                pg_table_actual=pg_map.get(key, ""),
            ))
    return bundles


def run_comparison():
    orapool = OraclePool(ORACLE_CONFIG, POOL_SIZE)
    pgpool = PostgresPool(POSTGRES_CONFIG, POOL_SIZE)
    bundles = []
    try:
        meta_workers = max(min(len(SCHEMA_PAIRS) * 12, POOL_SIZE * 2), 4)
        with cf.ThreadPoolExecutor(max_workers=meta_workers,
                                   thread_name_prefix="meta") as meta_ex:
            bundles = gather_metadata(orapool, pgpool, meta_ex)

        all_table_results = [r for b in bundles for r in b.table_results]
        log.info("Total tables (union): %d", len(all_table_results))

        with cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="ora") as ora_ex, \
             cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="pg") as pg_ex:
            futures = []
            for r in all_table_results:
                if r.in_oracle:    futures.append(ora_ex.submit(count_oracle, orapool, r))
                if r.in_postgres:  futures.append(pg_ex.submit(count_postgres, pgpool, r))
            done, total = 0, len(futures)
            for fut in cf.as_completed(futures):
                fut.result(); done += 1
                if done % 25 == 0 or done == total:
                    log.info("Progress: %d/%d count queries complete", done, total)

        # In-memory comparisons
        for b in bundles:
            common = ({t.lower() for t in b.ora_tables}
                      & {t.lower() for t in b.pg_tables})
            pg_actual_by_key = {t.lower(): t for t in b.pg_tables}

            if COMPARE_INDEXES:
                b.index_results = compare_indexes_for_pair(
                    b.ora_schema, b.pg_schema, common,
                    b.ora_indexes, b.pg_indexes)
            if COMPARE_COLUMNS:
                b.column_results = compare_columns_for_pair(
                    b.ora_schema, b.pg_schema, common,
                    b.ora_columns, b.pg_columns, pg_actual_by_key)
            if COMPARE_VIEWS:
                b.view_results = compare_simple_objects(
                    b.ora_schema, b.pg_schema, b.ora_views, b.pg_views)
            if COMPARE_MVIEWS:
                b.mview_results = compare_simple_objects(
                    b.ora_schema, b.pg_schema, b.ora_mviews, b.pg_mviews)
            if COMPARE_PROCS_AND_FUNCS:
                b.routine_results = compare_routines(
                    b.ora_schema, b.pg_schema, b.ora_routines, b.pg_routines)
            if COMPARE_TRIGGERS:
                b.trigger_results = compare_triggers(
                    b.ora_schema, b.pg_schema, b.ora_triggers, b.pg_triggers)
            if COMPARE_SEQUENCES:
                b.sequence_results = compare_sequences(
                    b.ora_schema, b.pg_schema, b.ora_sequences, b.pg_sequences)
    finally:
        orapool.close(); pgpool.close()

    return bundles


# ============================================================================
# XLSX report
# ============================================================================
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN  = PatternFill("solid", fgColor="C6EFCE")
RED    = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
ORANGE = PatternFill("solid", fgColor="FCD5B4")
THIN   = Border(*[Side(style="thin", color="D0D0D0")] * 4)
CENTER = Alignment(horizontal="center")


def _apply_headers(ws, headers, widths):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, CENTER, THIN
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def _match_fill(val):
    if val == "YES": return GREEN
    if val == "CROSS-KIND": return ORANGE
    if val == "ERROR": return YELLOW
    return RED


def write_row_count_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Exists",
               "Oracle_Count", "Postgres_Count", "Match", "Remarks"]
    _apply_headers(ws, headers, [16, 16, 34, 10, 14, 14, 10, 50])
    for i, r in enumerate(results, start=2):
        remarks = []
        if not r.in_oracle and r.in_postgres: remarks.append("Missing in Oracle")
        elif r.in_oracle and not r.in_postgres: remarks.append("Missing in Postgres")
        elif r.in_oracle and r.in_postgres \
                and r.oracle_count == 0 and r.pg_count == 0:
            remarks.append("Both sides empty (0 rows)")
        if r.oracle_error: remarks.append(f"Oracle: {r.oracle_error}")
        if r.pg_error:     remarks.append(f"Postgres: {r.pg_error}")
        ora_val = r.oracle_count if r.in_oracle and r.oracle_count is not None else "N/A"
        pg_val  = r.pg_count if r.in_postgres and r.pg_count is not None else "N/A"
        vals = [r.oracle_schema, r.pg_schema, r.table_name, r.exists_both,
                ora_val, pg_val, r.match, "; ".join(remarks)]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (4, 7): c.alignment = CENTER
            if col in (5, 6) and isinstance(v, int) and not isinstance(v, bool):
                c.number_format = "#,##0"
        ws.cell(row=i, column=7).fill = _match_fill(r.match)
        ws.cell(row=i, column=4).fill = GREEN if r.exists_both == "YES" else RED
    if results: ws.auto_filter.ref = f"A1:H{len(results) + 1}"


def write_column_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Column",
               "Oracle_Type", "Expected_PG_Type (SCT)", "Actual_PG_Type",
               "Deviation", "Match", "Remarks"]
    _apply_headers(ws, headers, [14, 14, 28, 22, 22, 25, 25, 18, 10, 42])
    for i, r in enumerate(results, start=2):
        vals = [r.oracle_schema, r.pg_schema, r.table_name, r.column_name,
                r.oracle_type, r.expected_pg_type, r.actual_pg_type,
                r.deviation, r.match, r.remarks]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (8, 9): c.alignment = CENTER
        ws.cell(row=i, column=9).fill = _match_fill(r.match)
        dev_cell = ws.cell(row=i, column=8)
        if r.deviation == "SCT_MATCH": dev_cell.fill = GREEN
        elif r.deviation == "DEVIATION": dev_cell.fill = ORANGE
        else: dev_cell.fill = RED
    if results: ws.auto_filter.ref = f"A1:J{len(results) + 1}"


def write_index_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Columns",
               "Oracle_Index", "Postgres_Index",
               "Oracle_Unique", "Postgres_Unique", "Match", "Remarks"]
    _apply_headers(ws, headers, [14, 14, 26, 36, 26, 26, 10, 10, 10, 40])
    for i, r in enumerate(results, start=2):
        vals = [r.oracle_schema, r.pg_schema, r.table_name, r.columns,
                r.oracle_index, r.pg_index, r.oracle_unique, r.pg_unique,
                r.match, r.remarks]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (7, 8, 9): c.alignment = CENTER
        ws.cell(row=i, column=9).fill = _match_fill(r.match)
    if results: ws.auto_filter.ref = f"A1:J{len(results) + 1}"


def write_simple_object_sheet(ws, results, name_label="Object_Name"):
    headers = ["Oracle_Schema", "Postgres_Schema", name_label,
               "In_Oracle", "In_Postgres", "Match", "Remarks"]
    _apply_headers(ws, headers, [16, 16, 34, 12, 12, 10, 50])
    for i, r in enumerate(results, start=2):
        vals = [r.oracle_schema, r.pg_schema, r.object_name,
                "YES" if r.in_oracle else "NO",
                "YES" if r.in_postgres else "NO",
                r.match, r.remarks]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (4, 5, 6): c.alignment = CENTER
        ws.cell(row=i, column=6).fill = _match_fill(r.match)
    if results: ws.auto_filter.ref = f"A1:G{len(results) + 1}"


def write_routine_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Routine_Name",
               "Oracle_Kind", "Postgres_Kind", "Match", "Remarks"]
    _apply_headers(ws, headers, [16, 16, 34, 16, 16, 14, 50])
    for i, r in enumerate(results, start=2):
        vals = [r.oracle_schema, r.pg_schema, r.routine_name,
                r.ora_kind, r.pg_kind, r.match, r.remarks]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (4, 5, 6): c.alignment = CENTER
        ws.cell(row=i, column=6).fill = _match_fill(r.match)
    if results: ws.auto_filter.ref = f"A1:G{len(results) + 1}"


def write_trigger_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Trigger_Name",
               "Oracle_Table", "Postgres_Table",
               "In_Oracle", "In_Postgres", "Match", "Remarks"]
    _apply_headers(ws, headers, [14, 14, 30, 24, 24, 10, 10, 10, 45])
    for i, r in enumerate(results, start=2):
        vals = [r.oracle_schema, r.pg_schema, r.trigger_name,
                r.ora_table, r.pg_table,
                "YES" if r.in_oracle else "NO",
                "YES" if r.in_postgres else "NO",
                r.match, r.remarks]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=col, value=v); c.border = THIN
            if col in (6, 7, 8): c.alignment = CENTER
        ws.cell(row=i, column=8).fill = _match_fill(r.match)
    if results: ws.auto_filter.ref = f"A1:I{len(results) + 1}"


def write_summary_sheet(ws, bundles, ddl_stats):
    def totals(results, cond):
        return sum(1 for r in results if cond(r))

    tr = [r for b in bundles for r in b.table_results]
    ir = [r for b in bundles for r in b.index_results]
    cr = [r for b in bundles for r in b.column_results]
    vr = [r for b in bundles for r in b.view_results]
    mr = [r for b in bundles for r in b.mview_results]
    rr = [r for b in bundles for r in b.routine_results]
    trg = [r for b in bundles for r in b.trigger_results]
    sqr = [r for b in bundles for r in b.sequence_results]
    ddl_gen, ddl_skip = ddl_stats or (0, 0)

    rows = [
        ("=== Tables ===", ""),
        ("Total tables (union)", len(tr)),
        ("Exists in both", totals(tr, lambda r: r.exists_both == "YES")),
        ("Missing in Postgres only", totals(tr, lambda r: r.in_oracle and not r.in_postgres)),
        ("Missing in Oracle only", totals(tr, lambda r: not r.in_oracle and r.in_postgres)),
        ("Row counts MATCH", totals(tr, lambda r: r.match == "YES")),
        ("Row counts MISMATCH", totals(tr, lambda r: r.match == "NO" and r.exists_both == "YES")),
        ("Errors", totals(tr, lambda r: r.match == "ERROR")),
        ("", ""),
        ("=== Columns (AWS SCT rules) ===", ""),
        ("Total column comparisons", len(cr)),
        ("SCT match", totals(cr, lambda r: r.deviation == "SCT_MATCH")),
        ("Deviation from SCT default", totals(cr, lambda r: r.deviation == "DEVIATION")),
        ("Missing in Postgres", totals(cr, lambda r: r.deviation == "MISSING_IN_PG")),
        ("Missing in Oracle", totals(cr, lambda r: r.deviation == "MISSING_IN_ORACLE")),
        ("Nullability differences", totals(cr, lambda r: "Nullability differs" in r.remarks)),
        ("", ""),
        ("=== Indexes ===", ""),
        ("Total index comparisons", len(ir)),
        ("Indexes matched", totals(ir, lambda r: r.match == "YES")),
        ("Missing in Postgres", totals(ir, lambda r: "Missing in Postgres" in r.remarks)),
        ("Missing in Oracle", totals(ir, lambda r: "Missing in Oracle" in r.remarks)),
        ("Uniqueness differs", totals(ir, lambda r: "Uniqueness differs" in r.remarks)),
        ("", ""),
        ("=== Views ===", ""),
        ("Total views", len(vr)),
        ("Matched", totals(vr, lambda r: r.match == "YES")),
        ("Missing in Postgres", totals(vr, lambda r: r.in_oracle and not r.in_postgres)),
        ("Missing in Oracle", totals(vr, lambda r: not r.in_oracle and r.in_postgres)),
        ("", ""),
        ("=== Materialized Views ===", ""),
        ("Total mviews", len(mr)),
        ("Matched", totals(mr, lambda r: r.match == "YES")),
        ("Missing in Postgres", totals(mr, lambda r: r.in_oracle and not r.in_postgres)),
        ("Missing in Oracle", totals(mr, lambda r: not r.in_oracle and r.in_postgres)),
        ("", ""),
        ("=== Procedures & Functions ===", ""),
        ("Total routines (union)", len(rr)),
        ("Matched (same kind)", totals(rr, lambda r: r.match == "YES")),
        ("Cross-kind (proc <-> function)", totals(rr, lambda r: r.match == "CROSS-KIND")),
        ("Missing in Postgres", totals(rr, lambda r: "Missing in Postgres" in r.remarks)),
        ("Missing in Oracle", totals(rr, lambda r: "Missing in Oracle" in r.remarks)),
        ("", ""),
        ("=== Triggers ===", ""),
        ("Total triggers", len(trg)),
        ("Matched", totals(trg, lambda r: r.match == "YES")),
        ("Missing in Postgres", totals(trg, lambda r: r.in_oracle and not r.in_postgres)),
        ("Missing in Oracle", totals(trg, lambda r: not r.in_oracle and r.in_postgres)),
        ("", ""),
        ("=== Sequences ===", ""),
        ("Total sequences", len(sqr)),
        ("Matched", totals(sqr, lambda r: r.match == "YES")),
        ("Missing in Postgres", totals(sqr, lambda r: r.in_oracle and not r.in_postgres)),
        ("Missing in Oracle", totals(sqr, lambda r: not r.in_oracle and r.in_postgres)),
        ("Implicit SERIAL sequences on PG side",
            totals(sqr, lambda r: "implicit" in r.remarks.lower())),
        ("", ""),
        ("=== Missing-index DDL ===", ""),
        ("CREATE INDEX statements generated", ddl_gen),
        ("Indexes skipped (manual translation needed)", ddl_skip),
        ("DDL file", MISSING_INDEX_DDL_FILE if GENERATE_MISSING_INDEX_DDL else "(disabled)"),
        ("", ""),
        ("Generated at", time.strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 24


def write_report(bundles, ddl_stats, path):
    wb = Workbook()

    ws = wb.active; ws.title = "Row Count"
    all_tr = sorted((r for b in bundles for r in b.table_results),
                    key=lambda r: (r.oracle_schema, r.table_name.lower()))
    write_row_count_sheet(ws, all_tr)

    if COMPARE_COLUMNS:
        ws = wb.create_sheet("Column Types")
        all_cr = sorted((r for b in bundles for r in b.column_results),
                        key=lambda r: (r.oracle_schema, r.table_name.lower(),
                                       r.column_name.lower()))
        write_column_sheet(ws, all_cr)

    if COMPARE_INDEXES:
        ws = wb.create_sheet("Indexes")
        all_ir = sorted((r for b in bundles for r in b.index_results),
                        key=lambda r: (r.oracle_schema, r.table_name, r.columns))
        write_index_sheet(ws, all_ir)

    if COMPARE_VIEWS:
        ws = wb.create_sheet("Views")
        write_simple_object_sheet(
            ws,
            sorted((r for b in bundles for r in b.view_results),
                   key=lambda r: (r.oracle_schema, r.object_name.lower())),
            name_label="View_Name")

    if COMPARE_MVIEWS:
        ws = wb.create_sheet("Materialized Views")
        write_simple_object_sheet(
            ws,
            sorted((r for b in bundles for r in b.mview_results),
                   key=lambda r: (r.oracle_schema, r.object_name.lower())),
            name_label="MView_Name")

    if COMPARE_PROCS_AND_FUNCS:
        ws = wb.create_sheet("Procedures & Functions")
        write_routine_sheet(
            ws,
            sorted((r for b in bundles for r in b.routine_results),
                   key=lambda r: (r.oracle_schema, r.routine_name.lower())))

    if COMPARE_TRIGGERS:
        ws = wb.create_sheet("Triggers")
        write_trigger_sheet(
            ws,
            sorted((r for b in bundles for r in b.trigger_results),
                   key=lambda r: (r.oracle_schema, r.trigger_name.lower())))

    if COMPARE_SEQUENCES:
        ws = wb.create_sheet("Sequences")
        write_simple_object_sheet(
            ws,
            sorted((r for b in bundles for r in b.sequence_results),
                   key=lambda r: (r.oracle_schema, r.object_name.lower())),
            name_label="Sequence_Name")

    ws = wb.create_sheet("Summary")
    write_summary_sheet(ws, bundles, ddl_stats)
    wb.save(path)
    log.info("Report written: %s", path)


# ============================================================================
# Main
# ============================================================================
def main():
    t0 = time.time()
    log.info("Starting: %d schema pair(s), %d workers/DB, pool size %d",
             len(SCHEMA_PAIRS), MAX_WORKERS_PER_DB, POOL_SIZE)
    bundles = run_comparison()

    ddl_stats = None
    if GENERATE_MISSING_INDEX_DDL and COMPARE_INDEXES:
        ddl_stats = generate_missing_pg_index_ddl(bundles, MISSING_INDEX_DDL_FILE)

    write_report(bundles, ddl_stats, OUTPUT_FILE)

    all_tr = [r for b in bundles for r in b.table_results]
    all_ir = [r for b in bundles for r in b.index_results]
    all_cr = [r for b in bundles for r in b.column_results]
    tbl_bad = [r for r in all_tr if r.match != "YES"]
    idx_bad = [r for r in all_ir if r.match != "YES"]
    col_bad = [r for r in all_cr if r.match != "YES"]

    log.info("Done in %.1fs.", time.time() - t0)
    log.info("Tables: %d/%d match | Indexes: %d/%d match | Columns: %d/%d SCT-clean",
             len(all_tr) - len(tbl_bad), len(all_tr),
             len(all_ir) - len(idx_bad), len(all_ir),
             len(all_cr) - len(col_bad), len(all_cr))
    if tbl_bad or idx_bad or col_bad:
        log.warning("Issues found -- see %s", OUTPUT_FILE)
        sys.exit(1)


if __name__ == "__main__":
    main()
