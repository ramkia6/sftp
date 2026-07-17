#!/usr/bin/env python3
"""
oracle_schema_compare.py
========================
Compare two Oracle databases schema-by-schema and produce an Excel diff report.

Compared categories
-------------------
  OBJECTS      ALL_OBJECTS inventory (type, name, status)
  TABLES       table-level attributes (partitioned, temporary, IOT)
  COLUMNS      ALL_TAB_COLUMNS: datatype, length, precision, scale, nullable, default, position
  INDEXES      ALL_INDEXES + column list, uniqueness, index type
  CONSTRAINTS  PK/UK/FK/CHECK keyed by (table, type, columns) so SYS_Cnnn names don't create noise
  SEQUENCES    increment, min/max, cycle, cache
  SOURCE       PL/SQL source (packages, procs, funcs, triggers, types) compared by normalized hash
  VIEWS        view text compared by normalized hash

Usage
-----
  Edit the CONFIG block below, then:  python oracle_schema_compare.py

Requires: python-oracledb, openpyxl
  pip install oracledb openpyxl
"""

from __future__ import annotations

import hashlib
import logging
import os
import re
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

try:
    import oracledb
except ImportError:  # pragma: no cover
    sys.exit("Missing dependency: pip install oracledb")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:  # pragma: no cover
    sys.exit("Missing dependency: pip install openpyxl")


LOG = logging.getLogger("ora_cmp")


# ═══════════════════════════════════════════════════════════════════════════ #
#  CONFIG  --  edit everything in this block, then just run the script.
# ═══════════════════════════════════════════════════════════════════════════ #
CONFIG = {
    # ---- source database -------------------------------------------------- #
    "source": {
        "user": "app_user",
        # Leave password as None to read the env var named in "password_env".
        "password": None,
        "password_env": "SRC_PWD",
        "dsn": "prod-host:1521/PRODPDB",
    },

    # ---- target database -------------------------------------------------- #
    "target": {
        "user": "app_user",
        "password": None,
        "password_env": "TGT_PWD",
        "dsn": "uat-host:1521/UATPDB",
    },

    # ---- schemas to compare ----------------------------------------------- #
    # List of (source_schema, target_schema). Use the same name twice when the
    # schema is named identically on both sides.
    "schema_pairs": [
        ("HR", "HR"),
        ("SALES", "SALES_UAT"),
    ],

    # ---- what to compare -------------------------------------------------- #
    # Any subset of: OBJECTS, TABLES, COLUMNS, INDEXES, CONSTRAINTS,
    #                SEQUENCES, SOURCE, VIEWS
    "categories": ["OBJECTS", "TABLES", "COLUMNS", "INDEXES",
                   "CONSTRAINTS", "SEQUENCES", "SOURCE", "VIEWS"],

    # Attribute names to skip entirely, e.g. ["STATUS", "CACHE_SIZE"].
    "ignore_attrs": [],

    # True  -> do not flag COLUMN_ID (ordinal position) differences.
    "ignore_column_order": False,

    # True  -> also write MATCH rows to the workbook (much larger file).
    "include_matches": False,

    # ---- output / runtime ------------------------------------------------- #
    # Set "output" to None for an auto-timestamped filename.
    "output": None,
    "threads": 8,
    "verbose": False,

    # Path to Oracle Instant Client for thick mode; None = thin mode (default).
    "thick_lib_dir": None,
}
# ═══════════════════════════════════════════════════════════════════════════ #


MISSING = "MISSING_IN_TARGET"
EXTRA = "EXTRA_IN_TARGET"
MISMATCH = "MISMATCH"
MATCH = "MATCH"

STATUS_ORDER = [MISSING, EXTRA, MISMATCH, MATCH]

FILLS = {
    MISSING: PatternFill("solid", fgColor="FFC7CE"),   # red
    EXTRA: PatternFill("solid", fgColor="FFE2B7"),     # orange
    MISMATCH: PatternFill("solid", fgColor="FFEB9C"),  # yellow
    MATCH: PatternFill("solid", fgColor="C6EFCE"),     # green
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


# --------------------------------------------------------------------------- #
# Metadata extraction specs
# --------------------------------------------------------------------------- #
@dataclass
class Spec:
    name: str                       # category / sheet name
    key_cols: List[str]             # columns forming the identity of a row
    attr_cols: List[str]            # columns compared as attributes
    sql: str                        # must bind :owner
    row_filter: Optional[Callable[[Dict[str, Any]], bool]] = None
    postprocess: Optional[Callable[[List[Dict[str, Any]]], List[Dict[str, Any]]]] = None
    enabled: bool = True


OBJECT_TYPES = (
    "'TABLE','VIEW','INDEX','SEQUENCE','PROCEDURE','FUNCTION','PACKAGE',"
    "'PACKAGE BODY','TRIGGER','TYPE','TYPE BODY','SYNONYM','MATERIALIZED VIEW'"
)

SQL_OBJECTS = f"""
SELECT o.object_type, o.object_name, o.status
  FROM all_objects o
 WHERE o.owner = :owner
   AND o.object_type IN ({OBJECT_TYPES})
   AND o.object_name NOT LIKE 'BIN$%'
   AND o.object_name NOT LIKE 'SYS\\_%' ESCAPE '\\'
"""

SQL_TABLES = """
SELECT t.table_name, t.partitioned, t.temporary, t.iot_type, t.cluster_name
  FROM all_tables t
 WHERE t.owner = :owner
   AND t.table_name NOT LIKE 'BIN$%'
   AND t.nested = 'NO'
   AND (t.iot_type IS NULL OR t.iot_type != 'IOT_OVERFLOW')
"""

SQL_COLUMNS = """
SELECT c.table_name,
       c.column_name,
       c.column_id,
       c.data_type,
       c.data_type_owner,
       c.data_length,
       c.data_precision,
       c.data_scale,
       c.char_length,
       c.char_used,
       c.nullable,
       c.data_default,
       c.virtual_column,
       c.identity_column
  FROM all_tab_columns c
 WHERE c.owner = :owner
   AND c.table_name NOT LIKE 'BIN$%'
   AND EXISTS (SELECT 1 FROM all_tables t
                WHERE t.owner = c.owner AND t.table_name = c.table_name)
"""

SQL_INDEXES = """
SELECT i.index_name,
       i.table_name,
       i.index_type,
       i.uniqueness,
       i.partitioned,
       (SELECT LISTAGG(ic.column_name || ' ' || ic.descend, ',')
                 WITHIN GROUP (ORDER BY ic.column_position)
          FROM all_ind_columns ic
         WHERE ic.index_owner = i.owner
           AND ic.index_name  = i.index_name) AS index_columns
  FROM all_indexes i
 WHERE i.owner = :owner
   AND i.index_name NOT LIKE 'BIN$%'
   AND i.table_name NOT LIKE 'BIN$%'
"""

SQL_CONSTRAINTS = """
SELECT c.table_name,
       c.constraint_type,
       c.constraint_name,
       c.status,
       c.deferrable,
       c.validated,
       c.delete_rule,
       c.search_condition,
       rc.table_name AS r_table_name,
       (SELECT LISTAGG(cc.column_name, ',') WITHIN GROUP (ORDER BY cc.position)
          FROM all_cons_columns cc
         WHERE cc.owner = c.owner
           AND cc.constraint_name = c.constraint_name) AS cons_columns
  FROM all_constraints c
  LEFT JOIN all_constraints rc
         ON rc.owner = c.r_owner
        AND rc.constraint_name = c.r_constraint_name
 WHERE c.owner = :owner
   AND c.constraint_type IN ('P', 'U', 'R', 'C')
   AND c.table_name NOT LIKE 'BIN$%'
"""

SQL_SEQUENCES = """
SELECT s.sequence_name, s.min_value, s.max_value, s.increment_by,
       s.cycle_flag, s.order_flag, s.cache_size
  FROM all_sequences s
 WHERE s.sequence_owner = :owner
"""

SQL_SOURCE = """
SELECT s.type, s.name, s.line, s.text
  FROM all_source s
 WHERE s.owner = :owner
   AND s.name NOT LIKE 'BIN$%'
 ORDER BY s.type, s.name, s.line
"""

SQL_VIEWS = """
SELECT v.view_name, v.text
  FROM all_views v
 WHERE v.owner = :owner
"""

NOT_NULL_CHECK = re.compile(r'^\s*"?[A-Za-z0-9_$#]+"?\s+IS\s+NOT\s+NULL\s*$', re.I)


def _drop_notnull_checks(row: Dict[str, Any]) -> bool:
    """Filter out the implicit CHECK constraints Oracle creates for NOT NULL."""
    if row.get("CONSTRAINT_TYPE") == "C":
        cond = row.get("SEARCH_CONDITION") or ""
        if NOT_NULL_CHECK.match(str(cond)):
            return False
    return True


def _hash_source(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Collapse ALL_SOURCE line rows into one hashed row per object."""
    buckets: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    for r in rows:
        buckets[(r["TYPE"], r["NAME"])].append(r["TEXT"] or "")
    out = []
    for (otype, name), lines in buckets.items():
        body = "".join(lines)
        out.append({
            "TYPE": otype,
            "NAME": name,
            "LINES": len(lines),
            "SOURCE_HASH": _norm_hash(body),
        })
    return out


def _hash_views(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out = []
    for r in rows:
        text = r.get("TEXT") or ""
        out.append({
            "VIEW_NAME": r["VIEW_NAME"],
            "TEXT_HASH": _norm_hash(text),
            "TEXT_LENGTH": len(text),
        })
    return out


def _norm_hash(text: str) -> str:
    """Whitespace-insensitive, case-insensitive-on-keywords-ish hash."""
    normalized = re.sub(r"\s+", " ", text or "").strip().upper()
    return hashlib.md5(normalized.encode("utf-8", "replace")).hexdigest()


def build_specs(cfg: dict) -> List[Spec]:
    specs = [
        Spec("OBJECTS", ["OBJECT_TYPE", "OBJECT_NAME"], ["STATUS"], SQL_OBJECTS),
        Spec("TABLES", ["TABLE_NAME"],
             ["PARTITIONED", "TEMPORARY", "IOT_TYPE", "CLUSTER_NAME"], SQL_TABLES),
        Spec("COLUMNS", ["TABLE_NAME", "COLUMN_NAME"],
             ["COLUMN_ID", "DATA_TYPE", "DATA_TYPE_OWNER", "DATA_LENGTH",
              "DATA_PRECISION", "DATA_SCALE", "CHAR_LENGTH", "CHAR_USED",
              "NULLABLE", "DATA_DEFAULT", "VIRTUAL_COLUMN", "IDENTITY_COLUMN"],
             SQL_COLUMNS),
        Spec("INDEXES", ["INDEX_NAME"],
             ["TABLE_NAME", "INDEX_TYPE", "UNIQUENESS", "PARTITIONED", "INDEX_COLUMNS"],
             SQL_INDEXES),
        Spec("CONSTRAINTS",
             ["TABLE_NAME", "CONSTRAINT_TYPE", "CONS_COLUMNS", "SEARCH_CONDITION"],
             ["CONSTRAINT_NAME", "STATUS", "VALIDATED", "DEFERRABLE",
              "DELETE_RULE", "R_TABLE_NAME"],
             SQL_CONSTRAINTS, row_filter=_drop_notnull_checks),
        Spec("SEQUENCES", ["SEQUENCE_NAME"],
             ["MIN_VALUE", "MAX_VALUE", "INCREMENT_BY", "CYCLE_FLAG",
              "ORDER_FLAG", "CACHE_SIZE"],
             SQL_SEQUENCES),
        Spec("SOURCE", ["TYPE", "NAME"], ["SOURCE_HASH", "LINES"],
             SQL_SOURCE, postprocess=_hash_source),
        Spec("VIEWS", ["VIEW_NAME"], ["TEXT_HASH", "TEXT_LENGTH"],
             SQL_VIEWS, postprocess=_hash_views),
    ]
    wanted = {c.strip().upper() for c in cfg.get("categories") or []}
    known = {s.name for s in specs}
    unknown = wanted - known
    if unknown:
        raise SystemExit(f"Unknown categories in CONFIG: {', '.join(sorted(unknown))}")
    if wanted:
        for s in specs:
            s.enabled = s.name in wanted
    if cfg.get("ignore_column_order"):
        for s in specs:
            if s.name == "COLUMNS" and "COLUMN_ID" in s.attr_cols:
                s.attr_cols.remove("COLUMN_ID")
    selected = [s for s in specs if s.enabled]
    if not selected:
        raise SystemExit("CONFIG['categories'] selects nothing to compare.")
    return selected


# --------------------------------------------------------------------------- #
# Extraction
# --------------------------------------------------------------------------- #
def make_pool(user: str, password: str, dsn: str, size: int) -> "oracledb.ConnectionPool":
    return oracledb.create_pool(
        user=user, password=password, dsn=dsn,
        min=1, max=max(size, 2), increment=1, timeout=60, getmode=oracledb.POOL_GETMODE_WAIT,
    )


def fetch(pool, spec: Spec, owner: str, side: str) -> Dict[Tuple, Dict[str, Any]]:
    """Run one extraction query and return {key_tuple: {attr: value}}."""
    t0 = time.time()
    with pool.acquire() as conn:
        cur = conn.cursor()
        cur.arraysize = 5000
        cur.prefetchrows = 5000
        cur.execute(spec.sql, owner=owner.upper())
        cols = [d[0].upper() for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        cur.close()

    if spec.row_filter:
        rows = [r for r in rows if spec.row_filter(r)]
    if spec.postprocess:
        rows = spec.postprocess(rows)

    result: Dict[Tuple, Dict[str, Any]] = {}
    for r in rows:
        key = tuple(norm(r.get(k)) for k in spec.key_cols)
        result[key] = {a: r.get(a) for a in spec.attr_cols}

    LOG.info("[%s] %-11s %-20s %6d rows  (%.1fs)",
             side, spec.name, owner.upper(), len(result), time.time() - t0)
    return result


def norm(v: Any) -> str:
    """Normalize a value for stable comparison / display."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


# --------------------------------------------------------------------------- #
# Comparison
# --------------------------------------------------------------------------- #
@dataclass
class DiffRow:
    src_schema: str
    tgt_schema: str
    key: Tuple
    status: str
    attribute: str
    src_value: str
    tgt_value: str


@dataclass
class CategoryResult:
    spec: Spec
    rows: List[DiffRow] = field(default_factory=list)
    counts: Dict[str, int] = field(default_factory=lambda: dict.fromkeys(STATUS_ORDER, 0))


def compare(spec: Spec, src: Dict, tgt: Dict, s_schema: str, t_schema: str,
            include_matches: bool, ignore_attrs: set) -> CategoryResult:
    res = CategoryResult(spec=spec)
    all_keys = sorted(set(src) | set(tgt))

    for key in all_keys:
        s, t = src.get(key), tgt.get(key)

        if s is not None and t is None:
            res.counts[MISSING] += 1
            res.rows.append(DiffRow(s_schema, t_schema, key, MISSING, "(object)",
                                    summarize(s), ""))
        elif s is None and t is not None:
            res.counts[EXTRA] += 1
            res.rows.append(DiffRow(s_schema, t_schema, key, EXTRA, "(object)",
                                    "", summarize(t)))
        else:
            deltas = [(a, norm(s[a]), norm(t[a]))
                      for a in spec.attr_cols
                      if a not in ignore_attrs and norm(s[a]) != norm(t[a])]
            if deltas:
                res.counts[MISMATCH] += 1
                for a, sv, tv in deltas:
                    res.rows.append(DiffRow(s_schema, t_schema, key, MISMATCH, a, sv, tv))
            else:
                res.counts[MATCH] += 1
                if include_matches:
                    res.rows.append(DiffRow(s_schema, t_schema, key, MATCH, "", "", ""))
    return res


def summarize(attrs: Dict[str, Any]) -> str:
    return ", ".join(f"{k}={norm(v)}" for k, v in attrs.items() if norm(v))[:32000]


# --------------------------------------------------------------------------- #
# Excel report
# --------------------------------------------------------------------------- #
def write_report(path: str, results: Dict[str, List[CategoryResult]], meta: Dict[str, str]):
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    _summary_sheet(ws, results, meta)

    for cat, cat_results in results.items():
        sheet = wb.create_sheet(cat[:31])
        spec = cat_results[0].spec
        headers = (["SOURCE_SCHEMA", "TARGET_SCHEMA"] + spec.key_cols +
                   ["STATUS", "ATTRIBUTE", "SOURCE_VALUE", "TARGET_VALUE"])
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill, cell.font = HEADER_FILL, HEADER_FONT
            cell.alignment = Alignment(vertical="center")

        for cr in cat_results:
            for r in cr.rows:
                sheet.append([r.src_schema, r.tgt_schema, *r.key,
                              r.status, r.attribute,
                              r.src_value[:32000], r.tgt_value[:32000]])
                fill = FILLS.get(r.status)
                if fill:
                    sheet.cell(row=sheet.max_row, column=len(spec.key_cols) + 3).fill = fill

        sheet.freeze_panes = "A2"
        if sheet.max_row > 1:
            sheet.auto_filter.ref = (
                f"A1:{get_column_letter(len(headers))}{sheet.max_row}")
        _autosize(sheet, len(headers))

    wb.save(path)
    LOG.info("Report written: %s", path)


def _summary_sheet(ws, results, meta):
    ws["A1"] = "Oracle Schema Comparison Report"
    ws["A1"].font = Font(size=14, bold=True)
    row = 3
    for k, v in meta.items():
        ws.cell(row=row, column=1, value=k).font = Font(bold=True)
        ws.cell(row=row, column=2, value=v)
        row += 1

    row += 1
    headers = ["CATEGORY", "SOURCE_SCHEMA", "TARGET_SCHEMA", *STATUS_ORDER, "TOTAL"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
    row += 1

    for cat, cat_results in results.items():
        for cr in cat_results:
            total = sum(cr.counts.values())
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=cr.rows[0].src_schema if cr.rows else "")
            ws.cell(row=row, column=3, value=cr.rows[0].tgt_schema if cr.rows else "")
            for i, st in enumerate(STATUS_ORDER, start=4):
                c = ws.cell(row=row, column=i, value=cr.counts[st])
                if cr.counts[st] and st != MATCH:
                    c.fill = FILLS[st]
            ws.cell(row=row, column=8, value=total)
            row += 1

    ws.freeze_panes = "A1"
    _autosize(ws, len(headers))


def _autosize(ws, ncols: int, cap: int = 60):
    for col in range(1, ncols + 1):
        width = 10
        for cell in ws[get_column_letter(col)][:400]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, cap))
        ws.column_dimensions[get_column_letter(col)].width = width


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def resolve_schema_pairs(cfg: dict) -> List[Tuple[str, str]]:
    pairs = cfg.get("schema_pairs") or []
    if not pairs:
        raise SystemExit("CONFIG['schema_pairs'] is empty.")
    out = []
    for item in pairs:
        if isinstance(item, str):          # allow a bare name for same-name schemas
            item = (item, item)
        if len(item) != 2:
            raise SystemExit(f"Bad schema_pairs entry: {item!r} (expected (SRC, TGT))")
        out.append((str(item[0]).strip().upper(), str(item[1]).strip().upper()))
    return out


def resolve_password(side_cfg: dict, label: str) -> str:
    pwd = side_cfg.get("password")
    if pwd:
        return pwd
    env = side_cfg.get("password_env")
    if env and os.getenv(env):
        return os.environ[env]
    raise SystemExit(
        f"No {label} password. Set CONFIG['{label}']['password'] "
        f"or export {env or '<password_env>'}."
    )


def main(cfg: dict = CONFIG) -> int:
    logging.basicConfig(
        level=logging.DEBUG if cfg.get("verbose") else logging.INFO,
        format="%(asctime)s %(levelname)-5s %(message)s", datefmt="%H:%M:%S")

    if cfg.get("thick_lib_dir"):
        oracledb.init_oracle_client(lib_dir=cfg["thick_lib_dir"])

    src_cfg, tgt_cfg = cfg["source"], cfg["target"]
    src_pwd = resolve_password(src_cfg, "source")
    tgt_pwd = resolve_password(tgt_cfg, "target")

    pairs = resolve_schema_pairs(cfg)
    specs = build_specs(cfg)
    ignore_attrs = {a.strip().upper() for a in cfg.get("ignore_attrs") or [] if a.strip()}
    threads = int(cfg.get("threads") or 8)
    include_matches = bool(cfg.get("include_matches"))
    out = cfg.get("output") or f"oracle_compare_{datetime.now():%Y%m%d_%H%M%S}.xlsx"

    LOG.info("Source      : %s@%s", src_cfg["user"], src_cfg["dsn"])
    LOG.info("Target      : %s@%s", tgt_cfg["user"], tgt_cfg["dsn"])
    LOG.info("Schema pairs: %s", ", ".join(f"{s}->{t}" for s, t in pairs))
    LOG.info("Categories  : %s", ", ".join(s.name for s in specs))

    t0 = time.time()
    src_pool = make_pool(src_cfg["user"], src_pwd, src_cfg["dsn"], threads)
    tgt_pool = make_pool(tgt_cfg["user"], tgt_pwd, tgt_cfg["dsn"], threads)

    try:
        # --- parallel metadata harvest -------------------------------------- #
        harvest: Dict[Tuple[str, str, str], Any] = {}
        with ThreadPoolExecutor(max_workers=threads) as ex:
            for spec in specs:
                for s_schema, t_schema in pairs:
                    harvest[("SRC", spec.name, s_schema)] = ex.submit(
                        fetch, src_pool, spec, s_schema, "SRC")
                    harvest[("TGT", spec.name, t_schema)] = ex.submit(
                        fetch, tgt_pool, spec, t_schema, "TGT")
            for f in as_completed(list(harvest.values())):
                f.result()  # surface exceptions early

        data = {k: f.result() for k, f in harvest.items()}
        LOG.info("Metadata harvested in %.1fs", time.time() - t0)
    finally:
        src_pool.close(force=True)
        tgt_pool.close(force=True)

    # --- compare ------------------------------------------------------------ #
    results: Dict[str, List[CategoryResult]] = {}
    for spec in specs:
        results[spec.name] = []
        for s_schema, t_schema in pairs:
            cr = compare(spec,
                         data[("SRC", spec.name, s_schema)],
                         data[("TGT", spec.name, t_schema)],
                         s_schema, t_schema, include_matches, ignore_attrs)
            results[spec.name].append(cr)
            LOG.info("%-11s %s->%s  missing=%d extra=%d mismatch=%d match=%d",
                     spec.name, s_schema, t_schema, cr.counts[MISSING],
                     cr.counts[EXTRA], cr.counts[MISMATCH], cr.counts[MATCH])

    meta = {
        "Generated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "Source": f'{src_cfg["user"]}@{src_cfg["dsn"]}',
        "Target": f'{tgt_cfg["user"]}@{tgt_cfg["dsn"]}',
        "Schema pairs": ", ".join(f"{s} -> {t}" for s, t in pairs),
        "Categories": ", ".join(s.name for s in specs),
        "Ignored attrs": ", ".join(sorted(ignore_attrs)) or "(none)",
        "Match rows included": str(include_matches),
    }
    write_report(out, results, meta)

    diffs = sum(cr.counts[MISSING] + cr.counts[EXTRA] + cr.counts[MISMATCH]
                for lst in results.values() for cr in lst)
    LOG.info("Done in %.1fs. Total differing objects: %d", time.time() - t0, diffs)
    return 1 if diffs else 0


if __name__ == "__main__":
    sys.exit(main())
