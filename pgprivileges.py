#!/usr/bin/env python3
"""
pg_privilege_audit.py
=====================
Audits a PostgreSQL / Aurora PostgreSQL instance and writes a clean Excel workbook
containing every role (group + user), the membership graph, and every privilege.

Sheets produced
---------------
  1. Summary                  - counts, connection info, run metadata
  2. Roles                    - all roles with attributes, classified GROUP / USER
  3. Group Members            - direct memberships (pg_auth_members)
  4. Membership Expanded      - recursive: user -> every role it reaches, with path
  5. Object Privileges        - every explicit GRANT on every object (long format)
  6. Default Privileges       - ALTER DEFAULT PRIVILEGES entries (pg_default_acl)
  7. Effective User Privileges- per login user: direct + inherited + PUBLIC grants

Usage (Windows)
---------------
  1. Edit the CONFIG block below - that's the only thing you need to touch.
  2. pip install psycopg2-binary openpyxl
  3. python pg_privilege_audit.py

Command-line flags still work and override CONFIG if you want them, e.g.
  python pg_privilege_audit.py --all-databases
  python pg_privilege_audit.py --host other-host --dbname otherdb

Requires: psycopg2-binary, openpyxl
"""

from __future__ import annotations

import argparse
import getpass
import os
import sys
from collections import defaultdict
from datetime import datetime

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# =========================================================================== #
#                        >>>  EDIT THIS BLOCK ONLY  <<<                        #
# =========================================================================== #
CONFIG = {
    # ---- connection ------------------------------------------------------- #
    "host":     "mydb.cluster-xxxxxxxx.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "user":     "admin",
    "dbname":   "postgres",

    # Password. Leave as None to read the PGPASSWORD environment variable, or
    # to be prompted securely at runtime. Only hardcode it if this file is NOT
    # going into source control (see SECURITY note at the bottom of this block).
    "password": None,

    # "require" for RDS/Aurora, "verify-full" if you supply a CA bundle below,
    # "disable" for a local trusted instance.
    "sslmode":  "require",
    "sslrootcert": None,          # e.g. r"C:\certs\us-east-1-bundle.pem"

    # ---- output ----------------------------------------------------------- #
    # Folder for the .xlsx. None = same folder as this script.
    # Use a raw string on Windows:  r"C:\Users\aditya\Desktop\audits"
    "output_dir":  None,
    # None = auto-name as pg_privileges_<host>_<timestamp>.xlsx
    "output_file": None,

    # ---- scope ------------------------------------------------------------ #
    # True  = loop every connectable database on the cluster into one workbook
    # False = just the "dbname" above
    "all_databases":    False,
    # True = include pg_* predefined roles and other system roles
    "include_system":   False,
    # True = also show implicit owner/PUBLIC defaults for objects with no
    #        explicit GRANT (complete, but much noisier)
    "include_implicit": False,

    # Open the workbook in Excel when finished (Windows only)
    "open_when_done": False,
}
# --------------------------------------------------------------------------- #
# SECURITY: if you set "password" above, this file now contains a live
# credential in plaintext. Do not commit it. Prefer one of:
#     setx PGPASSWORD "yourpassword"        (persists for new shells)
#     set  PGPASSWORD=yourpassword          (current shell only)
# ...and leave "password": None. With None and no env var, you get a hidden
# prompt at runtime, which is the safest option for an ad-hoc run.
# =========================================================================== #


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
LABEL_FONT = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MAX_COL_WIDTH = 55


# --------------------------------------------------------------------------- #
# SQL
# --------------------------------------------------------------------------- #
# grantee OID 0 is the pseudo-role PUBLIC
GRANTEE_EXPR = "CASE WHEN a.grantee = 0 THEN 'PUBLIC' ELSE pg_catalog.pg_get_userbyid(a.grantee) END"
GRANTOR_EXPR = "CASE WHEN a.grantor = 0 THEN '-' ELSE pg_catalog.pg_get_userbyid(a.grantor) END"

SQL_ROLES = """
SELECT r.rolname                                        AS role_name,
       CASE WHEN r.rolcanlogin THEN 'USER' ELSE 'GROUP' END AS role_type,
       r.rolsuper                                       AS superuser,
       r.rolinherit                                     AS inherit,
       r.rolcreaterole                                  AS create_role,
       r.rolcreatedb                                    AS create_db,
       r.rolcanlogin                                    AS can_login,
       r.rolreplication                                 AS replication,
       r.rolbypassrls                                   AS bypass_rls,
       CASE WHEN r.rolconnlimit < 0 THEN 'unlimited'
            ELSE r.rolconnlimit::text END               AS conn_limit,
       COALESCE(r.rolvaliduntil::text, 'infinity')      AS valid_until,
       (SELECT count(*) FROM pg_catalog.pg_auth_members m WHERE m.roleid = r.oid) AS direct_members,
       (SELECT count(*) FROM pg_catalog.pg_auth_members m WHERE m.member = r.oid) AS member_of_count,
       COALESCE(pg_catalog.shobj_description(r.oid, 'pg_authid'), '') AS comment
FROM pg_catalog.pg_roles r
WHERE {role_filter}
ORDER BY 2 DESC, 1;
"""

SQL_MEMBERS = """
SELECT g.rolname                                    AS group_name,
       m.rolname                                    AS member_name,
       CASE WHEN m.rolcanlogin THEN 'USER' ELSE 'GROUP' END AS member_type,
       am.admin_option                              AS admin_option,
       {extra_cols}
       COALESCE(gr.rolname, '-')                    AS granted_by
FROM pg_catalog.pg_auth_members am
JOIN pg_catalog.pg_roles g  ON g.oid  = am.roleid
JOIN pg_catalog.pg_roles m  ON m.oid  = am.member
LEFT JOIN pg_catalog.pg_roles gr ON gr.oid = am.grantor
WHERE {g_filter} AND {m_filter}
ORDER BY 1, 2;
"""

SQL_MEMBERS_EXPANDED = """
WITH RECURSIVE walk AS (
    SELECT am.member       AS start_oid,
           am.roleid       AS reached_oid,
           1               AS depth,
           ARRAY[am.member, am.roleid] AS path
    FROM pg_catalog.pg_auth_members am
    UNION ALL
    SELECT w.start_oid,
           am.roleid,
           w.depth + 1,
           w.path || am.roleid
    FROM walk w
    JOIN pg_catalog.pg_auth_members am ON am.member = w.reached_oid
    WHERE NOT am.roleid = ANY (w.path)
)
SELECT s.rolname                                          AS role_name,
       CASE WHEN s.rolcanlogin THEN 'USER' ELSE 'GROUP' END AS role_type,
       t.rolname                                          AS inherited_role,
       w.depth                                            AS depth,
       CASE WHEN w.depth = 1 THEN 'DIRECT' ELSE 'INDIRECT' END AS grant_kind,
       s.rolinherit                                       AS auto_inherits,
       (SELECT string_agg(p.rolname, ' -> ' ORDER BY u.ord)
          FROM unnest(w.path) WITH ORDINALITY AS u(oid, ord)
          JOIN pg_catalog.pg_roles p ON p.oid = u.oid)     AS membership_path
FROM walk w
JOIN pg_catalog.pg_roles s ON s.oid = w.start_oid
JOIN pg_catalog.pg_roles t ON t.oid = w.reached_oid
WHERE {s_filter} AND {t_filter}
ORDER BY 1, 4, 3;
"""

# --- object privileges: one query per catalog, unioned into a long-format sheet --
def sql_object_privs(include_implicit: bool) -> str:
    def acl(col, kind, owner):
        if include_implicit:
            return f"COALESCE({col}, pg_catalog.acldefault('{kind}', {owner}))"
        return col

    return f"""
-- databases -----------------------------------------------------------------
SELECT 'DATABASE'::text            AS object_type,
       ''::text                    AS schema_name,
       d.datname::text             AS object_name,
       ''::text                    AS column_name,
       pg_catalog.pg_get_userbyid(d.datdba)::text AS owner,
       {GRANTEE_EXPR}::text        AS grantee,
       a.privilege_type::text      AS privilege,
       a.is_grantable              AS grantable,
       {GRANTOR_EXPR}::text        AS grantor
FROM pg_catalog.pg_database d
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('d.datacl', 'd', 'd.datdba')}) a
WHERE d.datname = current_database()

UNION ALL
-- tablespaces ---------------------------------------------------------------
SELECT 'TABLESPACE', '', ts.spcname::text, '',
       pg_catalog.pg_get_userbyid(ts.spcowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_tablespace ts
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('ts.spcacl', 't', 'ts.spcowner')}) a

UNION ALL
-- schemas -------------------------------------------------------------------
SELECT 'SCHEMA', '', n.nspname::text, '',
       pg_catalog.pg_get_userbyid(n.nspowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_namespace n
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('n.nspacl', 'n', 'n.nspowner')}) a
WHERE n.nspname NOT LIKE 'pg\\_%' AND n.nspname <> 'information_schema'

UNION ALL
-- tables / views / matviews / foreign tables / sequences ---------------------
SELECT CASE c.relkind WHEN 'r' THEN 'TABLE'
                      WHEN 'p' THEN 'PARTITIONED TABLE'
                      WHEN 'v' THEN 'VIEW'
                      WHEN 'm' THEN 'MATERIALIZED VIEW'
                      WHEN 'f' THEN 'FOREIGN TABLE'
                      WHEN 'S' THEN 'SEQUENCE' END,
       n.nspname::text, c.relname::text, '',
       pg_catalog.pg_get_userbyid(c.relowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_class c
JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
CROSS JOIN LATERAL pg_catalog.aclexplode(
        {acl('c.relacl', 'r', 'c.relowner')}) a
WHERE c.relkind IN ('r','p','v','m','f','S')
  AND n.nspname NOT LIKE 'pg\\_%' AND n.nspname <> 'information_schema'

UNION ALL
-- column level grants -------------------------------------------------------
SELECT 'COLUMN', n.nspname::text, c.relname::text, att.attname::text,
       pg_catalog.pg_get_userbyid(c.relowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_attribute att
JOIN pg_catalog.pg_class c     ON c.oid = att.attrelid
JOIN pg_catalog.pg_namespace n ON n.oid = c.relnamespace
CROSS JOIN LATERAL pg_catalog.aclexplode(att.attacl) a
WHERE att.attacl IS NOT NULL
  AND att.attnum > 0 AND NOT att.attisdropped
  AND n.nspname NOT LIKE 'pg\\_%' AND n.nspname <> 'information_schema'

UNION ALL
-- functions / procedures ----------------------------------------------------
SELECT CASE p.prokind WHEN 'p' THEN 'PROCEDURE'
                      WHEN 'a' THEN 'AGGREGATE'
                      ELSE 'FUNCTION' END,
       n.nspname::text,
       (p.proname || '(' || pg_catalog.pg_get_function_identity_arguments(p.oid) || ')')::text,
       '',
       pg_catalog.pg_get_userbyid(p.proowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_proc p
JOIN pg_catalog.pg_namespace n ON n.oid = p.pronamespace
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('p.proacl', 'f', 'p.proowner')}) a
WHERE n.nspname NOT LIKE 'pg\\_%' AND n.nspname <> 'information_schema'

UNION ALL
-- languages -----------------------------------------------------------------
SELECT 'LANGUAGE', '', l.lanname::text, '',
       pg_catalog.pg_get_userbyid(l.lanowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_language l
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('l.lanacl', 'l', 'l.lanowner')}) a

UNION ALL
-- foreign servers -----------------------------------------------------------
SELECT 'FOREIGN SERVER', '', s.srvname::text, '',
       pg_catalog.pg_get_userbyid(s.srvowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_foreign_server s
CROSS JOIN LATERAL pg_catalog.aclexplode({acl('s.srvacl', 'S', 's.srvowner')}) a

UNION ALL
-- types / domains -----------------------------------------------------------
SELECT 'TYPE', n.nspname::text, t.typname::text, '',
       pg_catalog.pg_get_userbyid(t.typowner)::text,
       {GRANTEE_EXPR}::text, a.privilege_type::text, a.is_grantable, {GRANTOR_EXPR}::text
FROM pg_catalog.pg_type t
JOIN pg_catalog.pg_namespace n ON n.oid = t.typnamespace
CROSS JOIN LATERAL pg_catalog.aclexplode(t.typacl) a
WHERE t.typacl IS NOT NULL
  AND n.nspname NOT LIKE 'pg\\_%' AND n.nspname <> 'information_schema'

ORDER BY 1, 2, 3, 6, 7;
"""

SQL_DEFAULT_ACL = f"""
SELECT COALESCE(pg_catalog.pg_get_userbyid(da.defaclrole), '-')::text AS target_role,
       COALESCE(n.nspname, '<all schemas>')::text                     AS schema_name,
       CASE da.defaclobjtype WHEN 'r' THEN 'TABLES'
                             WHEN 'S' THEN 'SEQUENCES'
                             WHEN 'f' THEN 'FUNCTIONS'
                             WHEN 'T' THEN 'TYPES'
                             WHEN 'n' THEN 'SCHEMAS' END              AS object_type,
       {GRANTEE_EXPR}::text                                           AS grantee,
       a.privilege_type::text                                         AS privilege,
       a.is_grantable                                                 AS grantable
FROM pg_catalog.pg_default_acl da
LEFT JOIN pg_catalog.pg_namespace n ON n.oid = da.defaclnamespace
CROSS JOIN LATERAL pg_catalog.aclexplode(da.defaclacl) a
ORDER BY 1, 2, 3, 4, 5;
"""


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def role_filter(alias: str, include_system: bool) -> str:
    """
    Exclude only the pg_* predefined roles by default. Deliberately NOT filtering
    on oid > 16383: that would hide the bootstrap superuser (postgres, oid 10),
    which is the single most important role in a privilege audit.
    """
    if include_system:
        return "true"
    return f"{alias}.rolname NOT LIKE 'pg\\_%'"


def fetch(cur, sql: str) -> tuple[list[str], list[tuple]]:
    cur.execute(sql)
    cols = [d[0] for d in cur.description]
    return cols, cur.fetchall()


def norm(value):
    """Excel-friendly scalar."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Y" if value else "N"
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple],
                note: str | None = None) -> None:
    ws = wb.create_sheet(title[:31])
    start = 1

    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, color="808080", size=9)
        start = 2

    if not rows:
        ws.cell(row=start, column=1, value="(no rows)").font = Font(italic=True, color="808080")
        ws.column_dimensions["A"].width = 30
        return

    header_row = start
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h.replace("_", " ").title())
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for r, row in enumerate(rows, header_row + 1):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=norm(val))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top")

    last_row = header_row + len(rows)
    last_col = get_column_letter(len(headers))

    # widths
    for c, h in enumerate(headers, 1):
        longest = len(str(h))
        for row in rows:
            longest = max(longest, len(str(norm(row[c - 1]))))
        ws.column_dimensions[get_column_letter(c)].width = min(longest + 3, MAX_COL_WIDTH)

    # banded, filterable table
    ref = f"A{header_row}:{last_col}{last_row}"
    tbl = Table(displayName=f"tbl_{title.replace(' ', '_')[:25]}", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    try:
        ws.add_table(tbl)
    except ValueError:
        ws.auto_filter.ref = ref  # fallback if the name collides

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_summary(wb: Workbook, meta: list[tuple[str, object]]) -> None:
    ws = wb.create_sheet("Summary", 0)
    ws.cell(row=1, column=1, value="PostgreSQL Privilege Audit").font = TITLE_FONT
    r = 3
    for label, value in meta:
        if label == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=norm(value))
        r += 1
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70
    ws.sheet_view.showGridLines = False


# --------------------------------------------------------------------------- #
# Collectors
# --------------------------------------------------------------------------- #
def collect_effective(role_rows, member_rows, priv_rows, priv_cols):
    """
    Expand every login user's reachable roles and attach the object privileges
    granted to each of those roles (plus anything granted to PUBLIC).
    """
    role_type = {r[0]: r[1] for r in role_rows}
    inherits = {r[0]: r[3] for r in role_rows}          # rolinherit
    logins = [r[0] for r in role_rows if r[1] == "USER"]

    # role -> [(reached_role, depth, kind, path)]
    reach = defaultdict(list)
    for row in member_rows:
        reach[row[0]].append((row[2], row[3], row[4], row[6]))

    gi = priv_cols.index("grantee")
    by_grantee = defaultdict(list)
    for row in priv_rows:
        by_grantee[row[gi]].append(row)

    out = []
    for user in logins:
        sources = [(user, 0, "DIRECT", user)]
        sources += [(t[0], t[1], t[2], t[3]) for t in reach.get(user, [])]
        sources.append(("PUBLIC", 0, "PUBLIC", "PUBLIC"))

        for src, depth, kind, path in sources:
            for row in by_grantee.get(src, []):
                d = dict(zip(priv_cols, row))
                if src == user:
                    via, needs_set = "(direct grant)", "N"
                elif src == "PUBLIC":
                    via, needs_set = "PUBLIC", "N"
                else:
                    via = src
                    needs_set = "N" if inherits.get(user) else "Y"
                out.append((
                    user,
                    d["object_type"], d["schema_name"], d["object_name"], d["column_name"],
                    d["privilege"], d["grantable"],
                    kind if src != "PUBLIC" else "PUBLIC",
                    via, role_type.get(src, "-" if src == "PUBLIC" else ""),
                    path, needs_set,
                ))

    out.sort(key=lambda x: (x[0], x[1], x[2], x[3], x[5]))
    headers = ["user_name", "object_type", "schema_name", "object_name", "column_name",
               "privilege", "grantable", "grant_kind", "granted_via", "via_role_type",
               "membership_path", "needs_set_role"]
    return headers, out


def audit_database(conn, dbname, include_system, include_implicit):
    """Return per-database sheet payloads."""
    with conn.cursor() as cur:
        cur.execute("SHOW server_version_num;")
        ver = int(cur.fetchone()[0])

        # PG16+ exposes inherit_option / set_option on pg_auth_members
        extra = ""
        if ver >= 160000:
            extra = "am.inherit_option AS inherit_option, am.set_option AS set_option, "

        rf = role_filter("r", include_system)
        role_cols, role_rows = fetch(cur, SQL_ROLES.format(role_filter=rf))

        mem_cols, mem_rows = fetch(cur, SQL_MEMBERS.format(
            extra_cols=extra,
            g_filter=role_filter("g", include_system),
            m_filter=role_filter("m", include_system)))

        exp_cols, exp_rows = fetch(cur, SQL_MEMBERS_EXPANDED.format(
            s_filter=role_filter("s", include_system),
            t_filter=role_filter("t", include_system)))

        priv_cols, priv_rows = fetch(cur, sql_object_privs(include_implicit))
        dacl_cols, dacl_rows = fetch(cur, SQL_DEFAULT_ACL)

    if not include_system:
        keep = {r[0] for r in role_rows} | {"PUBLIC"}
        gi = priv_cols.index("grantee")
        priv_rows = [r for r in priv_rows if r[gi] in keep]
        dacl_rows = [r for r in dacl_rows if r[3] in keep]

    eff_cols, eff_rows = collect_effective(role_rows, exp_rows, priv_rows, priv_cols)

    return {
        "version": ver,
        "roles": (role_cols, role_rows),
        "members": (mem_cols, mem_rows),
        "expanded": (exp_cols, exp_rows),
        "privs": (priv_cols, priv_rows),
        "defacl": (dacl_cols, dacl_rows),
        "effective": (eff_cols, eff_rows),
    }


# --------------------------------------------------------------------------- #
# Main
# --------------------------------------------------------------------------- #
def main() -> int:
    p = argparse.ArgumentParser(
        description="Export PostgreSQL roles, members and privileges to Excel. "
                    "All settings default to the CONFIG block at the top of this file.")
    p.add_argument("--host", default=CONFIG["host"])
    p.add_argument("--port", type=int, default=CONFIG["port"])
    p.add_argument("--user", default=CONFIG["user"])
    p.add_argument("--dbname", default=CONFIG["dbname"])
    p.add_argument("--password", default=None)
    p.add_argument("--sslmode", default=CONFIG["sslmode"])
    p.add_argument("--out", default=None, help="full path to the output .xlsx")
    p.add_argument("--all-databases", action="store_true", default=CONFIG["all_databases"],
                   help="loop over every connectable database; object sheets get a database column")
    p.add_argument("--include-system", action="store_true", default=CONFIG["include_system"],
                   help="include pg_* predefined roles and other system roles")
    p.add_argument("--include-implicit", action="store_true", default=CONFIG["include_implicit"],
                   help="also show implicit owner/PUBLIC defaults for objects with no explicit GRANT")
    args = p.parse_args()

    # password precedence: --password > CONFIG > PGPASSWORD env > secure prompt
    pwd = args.password or CONFIG["password"] or os.getenv("PGPASSWORD")
    if not pwd:
        pwd = getpass.getpass(f"Password for {args.user}@{args.host}: ")

    base = dict(host=args.host, port=args.port, user=args.user,
                password=pwd, sslmode=args.sslmode, connect_timeout=15)
    if CONFIG.get("sslrootcert"):
        base["sslrootcert"] = CONFIG["sslrootcert"]

    ts = datetime.now()
    if args.out:
        out = args.out
    else:
        fname = CONFIG["output_file"] or f"pg_privileges_{args.host.split('.')[0]}_{ts:%Y%m%d_%H%M%S}.xlsx"
        folder = CONFIG["output_dir"] or os.path.dirname(os.path.abspath(__file__))
        os.makedirs(folder, exist_ok=True)
        out = os.path.join(folder, fname)

    print(f"Connecting to {args.user}@{args.host}:{args.port}/{args.dbname} (sslmode={args.sslmode}) ...")

    # which databases?
    try:
        conn = psycopg2.connect(dbname=args.dbname, **base)
    except psycopg2.Error as e:
        print(f"ERROR: cannot connect: {e}", file=sys.stderr)
        return 2
    conn.set_session(readonly=True, autocommit=True)

    dbs = [args.dbname]
    if args.all_databases:
        with conn.cursor() as cur:
            cur.execute("SELECT datname FROM pg_database "
                        "WHERE datallowconn AND NOT datistemplate ORDER BY 1;")
            dbs = [r[0] for r in cur.fetchall()]

    results: dict[str, dict] = {}
    for db in dbs:
        c = conn if db == args.dbname else psycopg2.connect(dbname=db, **base)
        if c is not conn:
            c.set_session(readonly=True, autocommit=True)
        try:
            print(f"  scanning {db} ...", flush=True)
            results[db] = audit_database(c, db, args.include_system, args.include_implicit)
        except psycopg2.Error as e:
            print(f"  WARNING: skipping {db}: {str(e).strip()}", file=sys.stderr)
        finally:
            if c is not conn:
                c.close()

    if not results:
        print("ERROR: nothing collected.", file=sys.stderr)
        conn.close()
        return 2

    first = results[dbs[0] if dbs[0] in results else next(iter(results))]

    # roles + memberships are cluster-wide -> take them from the first database
    wb = Workbook()
    wb.remove(wb.active)

    def merge(key, add_db_col):
        cols = list(first[key][0])
        rows = []
        if add_db_col and len(results) > 1:
            cols = ["database"] + cols
            for db, res in results.items():
                rows += [(db,) + tuple(r) for r in res[key][1]]
        else:
            rows = [tuple(r) for r in first[key][1]]
        return cols, rows

    role_cols, role_rows = first["roles"]
    mem_cols, mem_rows = first["members"]
    exp_cols, exp_rows = first["expanded"]
    priv_cols, priv_rows = merge("privs", True)
    dacl_cols, dacl_rows = merge("defacl", True)
    eff_cols, eff_rows = merge("effective", True)

    n_groups = sum(1 for r in role_rows if r[1] == "GROUP")
    n_users = sum(1 for r in role_rows if r[1] == "USER")
    n_super = sum(1 for r in role_rows if r[2])

    write_summary(wb, [
        ("Generated at", ts.strftime("%Y-%m-%d %H:%M:%S")),
        ("Host", f"{args.host}:{args.port}"),
        ("Connected as", args.user),
        ("Databases audited", ", ".join(results.keys())),
        ("Server version", first["version"]),
        ("System roles included", "Yes" if args.include_system else "No"),
        ("Implicit defaults included", "Yes" if args.include_implicit else "No"),
        ("", ""),
        ("Roles total", len(role_rows)),
        ("  Groups (NOLOGIN)", n_groups),
        ("  Users (LOGIN)", n_users),
        ("  Superusers", n_super),
        ("Direct memberships", len(mem_rows)),
        ("Expanded memberships", len(exp_rows)),
        ("Explicit object grants", len(priv_rows)),
        ("Default privilege entries", len(dacl_rows)),
        ("Effective user privileges", len(eff_rows)),
    ])

    write_sheet(wb, "Roles", role_cols, role_rows,
                note="Every role in the cluster. Roles are cluster-wide; GROUP = NOLOGIN, USER = LOGIN.")
    write_sheet(wb, "Group Members", mem_cols, mem_rows,
                note="Direct GRANT <group> TO <member> edges only. See 'Membership Expanded' for the full closure.")
    write_sheet(wb, "Membership Expanded", exp_cols, exp_rows,
                note="Recursive closure. auto_inherits = N means the role must SET ROLE to use the inherited privileges.")
    write_sheet(wb, "Object Privileges", priv_cols, priv_rows,
                note="Explicit GRANTs per object. grantee PUBLIC = granted to everyone. Object privileges are per-database.")
    write_sheet(wb, "Default Privileges", dacl_cols, dacl_rows,
                note="ALTER DEFAULT PRIVILEGES entries - applied to objects created in future by target_role.")
    write_sheet(wb, "Effective User Privileges", eff_cols, eff_rows,
                note="Per login user: direct grants + everything inherited through groups + PUBLIC. Owner rights are implicit and only shown with --include-implicit.")

    try:
        wb.save(out)
    except PermissionError:
        conn.close()
        print(f"\nERROR: cannot write {out} - the file is open in Excel. "
              f"Close it and re-run.", file=sys.stderr)
        return 3

    conn.close()
    print(f"\nWrote {out}")
    print(f"  {len(role_rows)} roles ({n_groups} groups / {n_users} users), "
          f"{len(priv_rows)} explicit grants, {len(eff_rows)} effective privilege rows")

    if CONFIG.get("open_when_done") and os.name == "nt":
        os.startfile(out)  # noqa: S606  (Windows only)
    return 0


if __name__ == "__main__":
    sys.exit(main())
