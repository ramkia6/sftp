#!/usr/bin/env python3
"""
Multi-threaded Oracle vs PostgreSQL (Aurora) row-count comparison tool.
Oracle-driven. Detailed file logging captures every SQL + every DB response
to help debug "missing" tables.

Log file: db_rowcount_compare.log  (DEBUG level, fully verbose)
Console:  INFO level

Requirements:
    pip install oracledb psycopg2-binary openpyxl
"""

import concurrent.futures as cf
import logging
import sys
import time
from dataclasses import dataclass
from typing import Optional

import oracledb
import psycopg2
from psycopg2 import pool as pg_pool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIG -- edit these
# ----------------------------------------------------------------------------
ORACLE_CONFIG = {
    "user":     "oracle_user",
    "password": "oracle_password",
    "dsn":      "oracle-host.example.com:1521/ORCLPDB1",
}

POSTGRES_CONFIG = {
    "user":     "pg_user",
    "password": "pg_password",
    "host":     "aurora-cluster.cluster-xxxx.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "mydb",
}

SCHEMAS = ["SALES", "HR", "FINANCE"]
SCHEMA_OVERRIDES: dict = {
    # "ORACLE_NAME": "postgres_name",
}

MAX_WORKERS_PER_DB = 16
POOL_SIZE = MAX_WORKERS_PER_DB
OUTPUT_FILE = "row_count_comparison_report.xlsx"
LOG_FILE = "db_rowcount_compare.log"

# ----------------------------------------------------------------------------
# Logging: console at INFO, file at DEBUG (full SQL + responses)
# ----------------------------------------------------------------------------
log = logging.getLogger("rowcount")
log.setLevel(logging.DEBUG)
log.propagate = False

_fmt = logging.Formatter(
    "%(asctime)s [%(threadName)-12s] %(levelname)-7s %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)

_console = logging.StreamHandler(sys.stdout)
_console.setLevel(logging.INFO)
_console.setFormatter(_fmt)
log.addHandler(_console)

_file = logging.FileHandler(LOG_FILE, mode="w", encoding="utf-8")
_file.setLevel(logging.DEBUG)
_file.setFormatter(_fmt)
log.addHandler(_file)


def pg_schema_for(ora_schema: str) -> str:
    return SCHEMA_OVERRIDES.get(ora_schema, ora_schema.lower())


# ----------------------------------------------------------------------------
# Result model
# ----------------------------------------------------------------------------
@dataclass
class TableResult:
    oracle_schema: str
    pg_schema: str
    table_name: str
    ora_table_actual: str = ""
    pg_table_actual: str = ""
    in_oracle: bool = True
    in_postgres: bool = True
    oracle_count: Optional[int] = None
    pg_count: Optional[int] = None
    oracle_error: str = ""
    pg_error: str = ""

    @property
    def exists_both(self) -> str:
        return "YES" if (self.in_oracle and self.in_postgres) else "NO"

    @property
    def match(self) -> str:
        if not (self.in_oracle and self.in_postgres):
            return "NO"
        if self.oracle_count is None or self.pg_count is None:
            return "ERROR"
        return "YES" if self.oracle_count == self.pg_count else "NO"


# ----------------------------------------------------------------------------
# Connection pools
# ----------------------------------------------------------------------------
class OraclePool:
    def __init__(self, cfg, size):
        self.cfg = cfg
        self.pool = oracledb.create_pool(
            user=cfg["user"], password=cfg["password"], dsn=cfg["dsn"],
            min=2, max=size, increment=1,
            getmode=oracledb.POOL_GETMODE_WAIT,
        )

    def query_one(self, sql, params=None):
        conn = self.pool.acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchone()
        finally:
            self.pool.release(conn)

    def query_all(self, sql, params=None):
        conn = self.pool.acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchall()
        finally:
            self.pool.release(conn)

    def identity(self) -> str:
        return f"oracle user={self.cfg['user']} dsn={self.cfg['dsn']}"

    def close(self):
        self.pool.close(force=True)


class PostgresPool:
    def __init__(self, cfg, size):
        self.cfg = cfg
        self.pool = pg_pool.ThreadedConnectionPool(minconn=2, maxconn=size, **cfg)

    def query_one(self, sql, params=None):
        conn = self.pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                return cur.fetchone()
        finally:
            conn.rollback()
            self.pool.putconn(conn)

    def query_all(self, sql, params=None):
        conn = self.pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                return cur.fetchall()
        finally:
            conn.rollback()
            self.pool.putconn(conn)

    def identity(self) -> str:
        return (f"postgres user={self.cfg['user']} db={self.cfg['dbname']} "
                f"host={self.cfg['host']}:{self.cfg['port']}")

    def close(self):
        self.pool.closeall()


# ----------------------------------------------------------------------------
# Error formatters -- pull EVERYTHING useful out of the DB exception
# ----------------------------------------------------------------------------
def format_pg_error(e: BaseException) -> str:
    """Build a multi-line diagnostic from a psycopg2 exception."""
    lines = [f"  exception_type: {type(e).__name__}"]
    if isinstance(e, psycopg2.Error):
        pgcode = getattr(e, "pgcode", None)
        pgerror = getattr(e, "pgerror", None)
        lines.append(f"  pgcode (SQLSTATE): {pgcode}")
        if pgerror:
            lines.append(f"  pgerror: {pgerror.strip()}")
        diag = getattr(e, "diag", None)
        if diag is not None:
            # Dump every diag field that has a value
            for attr in ("severity", "sqlstate", "message_primary",
                         "message_detail", "message_hint", "statement_position",
                         "internal_position", "internal_query", "context",
                         "schema_name", "table_name", "column_name",
                         "datatype_name", "constraint_name", "source_file",
                         "source_line", "source_function"):
                val = getattr(diag, attr, None)
                if val:
                    lines.append(f"  diag.{attr}: {val}")
    lines.append(f"  str(e): {str(e).strip()}")
    return "\n".join(lines)


def format_ora_error(e: BaseException) -> str:
    lines = [f"  exception_type: {type(e).__name__}"]
    if isinstance(e, oracledb.DatabaseError) and e.args:
        err = e.args[0]
        lines.append(f"  code: ORA-{getattr(err, 'code', '?'):05d}")
        msg = getattr(err, "message", None)
        if msg:
            lines.append(f"  message: {msg.strip()}")
        ctx = getattr(err, "context", None)
        if ctx:
            lines.append(f"  context: {ctx}")
    lines.append(f"  str(e): {str(e).strip()}")
    return "\n".join(lines)


# ----------------------------------------------------------------------------
# Diagnostic: when a table is "missing" in Postgres, find similar names
# ----------------------------------------------------------------------------
def diagnose_missing_pg(pgpool: PostgresPool, schema: str, attempted: str) -> str:
    """
    On 42P01, query pg_class for any relation in this schema whose name
    matches case-insensitively. This catches the #1 cause: mixed-case
    identifiers in Postgres ("Employees" vs employees).

    Also reports whether the schema itself exists.
    """
    lines = []
    try:
        schema_check = pgpool.query_one(
            "SELECT 1 FROM pg_namespace WHERE nspname = %s", (schema,))
        if not schema_check:
            lines.append(f"  DIAGNOSIS: schema '{schema}' does NOT exist in pg_namespace")
            # Find similar schema names
            similar = pgpool.query_all(
                "SELECT nspname FROM pg_namespace WHERE lower(nspname) = lower(%s) "
                "OR nspname ILIKE %s LIMIT 10",
                (schema, f"%{schema}%"))
            if similar:
                lines.append(f"  similar schema names: {[s[0] for s in similar]}")
            return "\n".join(lines)
        lines.append(f"  schema '{schema}' exists in pg_namespace")
    except Exception as e:
        lines.append(f"  (schema existence check failed: {e})")
        return "\n".join(lines)

    relkind_map = {
        'r': 'ordinary table', 'p': 'partitioned table', 'f': 'foreign table',
        'v': 'view', 'm': 'materialized view', 'i': 'index', 'S': 'sequence',
        't': 'TOAST table', 'c': 'composite type',
    }
    try:
        rows = pgpool.query_all(
            """
            SELECT c.relname,
                   c.relkind,
                   pg_catalog.pg_get_userbyid(c.relowner)        AS owner,
                   has_table_privilege(current_user,
                                       quote_ident(n.nspname) || '.' ||
                                       quote_ident(c.relname),
                                       'SELECT')                AS can_select
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE n.nspname = %s
              AND lower(c.relname) = lower(%s)
            """,
            (schema, attempted),
        )
        if not rows:
            lines.append(f"  DIAGNOSIS: no relation in '{schema}' matches "
                         f"'{attempted}' (case-insensitive). Table truly absent.")
        else:
            lines.append(f"  DIAGNOSIS: found {len(rows)} relation(s) matching "
                         f"case-insensitively -- likely a case-sensitivity issue:")
            for relname, kind, owner, can_select in rows:
                kind_desc = relkind_map.get(kind, f"unknown ({kind})")
                lines.append(
                    f"    - relname='{relname}'  relkind='{kind}' ({kind_desc})  "
                    f"owner='{owner}'  can_select={can_select}"
                )
                if relname != attempted:
                    lines.append(
                        f"      -> case mismatch: tool tried \"{attempted}\", "
                        f"actual is \"{relname}\". The Postgres table was "
                        f"likely created with double-quoted mixed-case identifier."
                    )
                if not can_select:
                    lines.append(f"      -> current user lacks SELECT privilege")
    except Exception as e:
        lines.append(f"  (pg_class diagnostic query failed: {e})")
    return "\n".join(lines)


# ----------------------------------------------------------------------------
# Oracle metadata
# ----------------------------------------------------------------------------
def get_oracle_tables(orapool, schema):
    sql = """
        SELECT table_name
        FROM all_tables
        WHERE owner = :owner
          AND table_name NOT LIKE 'BIN$%'
          AND (nested = 'NO' OR nested IS NULL)
          AND iot_type IS NULL
    """
    log.debug("Oracle metadata SQL for schema %s:\n%s\n  params: owner=%s",
              schema, sql.strip(), schema.upper())
    rows = orapool.query_all(sql, {"owner": schema.upper()})
    tables = sorted(r[0] for r in rows)
    log.info("Oracle schema %s: %d tables found", schema, len(tables))
    log.debug("Oracle schema %s tables: %s", schema, tables)
    return tables


# ----------------------------------------------------------------------------
# Count workers
# ----------------------------------------------------------------------------
PG_UNDEFINED_TABLE = "42P01"
PG_UNDEFINED_SCHEMA = "3F000"
ORA_TABLE_OR_VIEW_NOT_EXIST = 942


def count_oracle(orapool, result: TableResult):
    sql = (f'SELECT COUNT(*) FROM '
           f'"{result.oracle_schema.upper()}"."{result.ora_table_actual}"')
    log.debug("[Oracle SQL]  %s.%s\n  %s\n  connection: %s",
              result.oracle_schema, result.ora_table_actual,
              sql, orapool.identity())
    t0 = time.time()
    try:
        row = orapool.query_one(sql)
        result.oracle_count = int(row[0])
        log.info("Oracle  %s.%s = %s (%.1fs)",
                 result.oracle_schema, result.ora_table_actual,
                 f"{result.oracle_count:,}", time.time() - t0)
        log.debug("[Oracle RESPONSE]  %s.%s -> %s",
                  result.oracle_schema, result.ora_table_actual, result.oracle_count)
    except oracledb.DatabaseError as e:
        err_details = format_ora_error(e)
        err, = e.args
        if getattr(err, "code", None) == ORA_TABLE_OR_VIEW_NOT_EXIST:
            result.in_oracle = False
            result.oracle_error = "Table not found at count time (ORA-00942)"
        else:
            result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle COUNT failed for %s.%s\n  SQL: %s\n  connection: %s\n%s",
                  result.oracle_schema, result.ora_table_actual,
                  sql, orapool.identity(), err_details)
    except Exception as e:
        result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle COUNT failed for %s.%s\n  SQL: %s\n  connection: %s\n%s",
                  result.oracle_schema, result.ora_table_actual,
                  sql, orapool.identity(), format_ora_error(e))


def count_postgres(pgpool, result: TableResult):
    sql = (f'SELECT COUNT(*) FROM '
           f'"{result.pg_schema}"."{result.pg_table_actual}"')
    log.debug("[Postgres SQL]  %s.%s\n  %s\n  connection: %s",
              result.pg_schema, result.pg_table_actual,
              sql, pgpool.identity())
    t0 = time.time()
    try:
        row = pgpool.query_one(sql)
        result.pg_count = int(row[0])
        log.info("Postgres %s.%s = %s (%.1fs)",
                 result.pg_schema, result.pg_table_actual,
                 f"{result.pg_count:,}", time.time() - t0)
        log.debug("[Postgres RESPONSE]  %s.%s -> %s",
                  result.pg_schema, result.pg_table_actual, result.pg_count)
    except psycopg2.Error as e:
        code = getattr(e, "pgcode", None)
        err_details = format_pg_error(e)

        # Header for the failure block in the log
        log.warning(
            "Postgres COUNT failed for %s.%s\n"
            "  SQL: %s\n"
            "  connection: %s\n"
            "%s",
            result.pg_schema, result.pg_table_actual,
            sql, pgpool.identity(), err_details,
        )

        if code in (PG_UNDEFINED_TABLE, PG_UNDEFINED_SCHEMA):
            result.in_postgres = False
            result.pg_error = ("Schema not found (3F000)"
                               if code == PG_UNDEFINED_SCHEMA
                               else "Table not found (42P01)")
            # Run the diagnostic and log it
            diag = diagnose_missing_pg(pgpool, result.pg_schema,
                                       result.pg_table_actual)
            log.warning("Diagnostic for missing %s.%s:\n%s",
                        result.pg_schema, result.pg_table_actual, diag)
        else:
            result.pg_error = f"{code}: {str(e).splitlines()[0][:180]}"
    except Exception as e:
        result.pg_error = str(e).splitlines()[0][:200]
        log.error("Postgres COUNT failed for %s.%s\n  SQL: %s\n  connection: %s\n%s",
                  result.pg_schema, result.pg_table_actual,
                  sql, pgpool.identity(), format_pg_error(e))


# ----------------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------------
def run_comparison():
    orapool = OraclePool(ORACLE_CONFIG, POOL_SIZE)
    pgpool = PostgresPool(POSTGRES_CONFIG, POOL_SIZE)
    log.info("Connected to %s", orapool.identity())
    log.info("Connected to %s", pgpool.identity())
    all_results = []

    try:
        with cf.ThreadPoolExecutor(max_workers=max(len(SCHEMAS), 2),
                                   thread_name_prefix="meta") as meta_ex:
            futs = {meta_ex.submit(get_oracle_tables, orapool, s): s
                    for s in SCHEMAS}
            for fut in cf.as_completed(futs):
                schema = futs[fut]
                pg_schema = pg_schema_for(schema)
                for tname in fut.result():
                    all_results.append(TableResult(
                        oracle_schema=schema,
                        pg_schema=pg_schema,
                        table_name=tname,
                        ora_table_actual=tname,
                        pg_table_actual=tname.lower(),
                    ))

        log.info("Total tables to process: %d", len(all_results))

        with cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="ora") as ora_ex, \
             cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="pg") as pg_ex:

            futures = []
            for r in all_results:
                futures.append(ora_ex.submit(count_oracle, orapool, r))
                futures.append(pg_ex.submit(count_postgres, pgpool, r))

            done, total = 0, len(futures)
            for fut in cf.as_completed(futures):
                fut.result()
                done += 1
                if done % 25 == 0 or done == total:
                    log.info("Progress: %d/%d count queries complete", done, total)
    finally:
        orapool.close()
        pgpool.close()

    all_results.sort(key=lambda r: (r.oracle_schema, r.table_name.lower()))
    return all_results


# ----------------------------------------------------------------------------
# XLSX report
# ----------------------------------------------------------------------------
def write_report(results, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Row Count Comparison"

    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Exists",
               "Oracle_Count", "Postgres_Count", "Match", "Remarks"]

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    thin = Border(*[Side(style="thin", color="D0D0D0")] * 4)
    center = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = header_fill, header_font, center, thin

    for i, r in enumerate(results, start=2):
        remarks = []
        if r.in_oracle and not r.in_postgres:
            remarks.append(f"Missing in Postgres (see {LOG_FILE} for diagnosis)")
        elif (r.in_oracle and r.in_postgres
              and r.oracle_count == 0 and r.pg_count == 0):
            remarks.append("Both sides empty (0 rows)")
        if r.oracle_error:
            remarks.append(f"Oracle: {r.oracle_error}")
        if r.pg_error and r.in_postgres:
            remarks.append(f"Postgres: {r.pg_error}")

        ora_val = (r.oracle_count if r.in_oracle and r.oracle_count is not None
                   else "N/A")
        pg_val = (r.pg_count if r.in_postgres and r.pg_count is not None
                  else "N/A")

        row_vals = [r.oracle_schema, r.pg_schema, r.table_name,
                    r.exists_both, ora_val, pg_val, r.match,
                    "; ".join(remarks)]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = thin
            if col in (4, 7):
                c.alignment = center
            if col in (5, 6) and isinstance(v, int) and not isinstance(v, bool):
                c.number_format = "#,##0"

        match_cell = ws.cell(row=i, column=7)
        exists_cell = ws.cell(row=i, column=4)
        match_cell.fill = (green if r.match == "YES"
                           else yellow if r.match == "ERROR" else red)
        exists_cell.fill = green if r.exists_both == "YES" else red

    widths = [18, 18, 38, 10, 16, 16, 10, 60]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"
    ws.freeze_panes = "A2"

    s = wb.create_sheet("Summary")
    total = len(results)
    both = sum(1 for r in results if r.exists_both == "YES")
    matched = sum(1 for r in results if r.match == "YES")
    mismatched = sum(1 for r in results
                     if r.match == "NO" and r.exists_both == "YES")
    missing_pg = sum(1 for r in results if r.in_oracle and not r.in_postgres)
    errors = sum(1 for r in results if r.match == "ERROR")
    both_empty = sum(1 for r in results
                     if r.in_oracle and r.in_postgres
                     and r.oracle_count == 0 and r.pg_count == 0)
    rows = [
        ("Total Oracle tables processed", total),
        ("Exists in both", both),
        ("Missing in Postgres", missing_pg),
        ("Row counts MATCH", matched),
        ("  ...of which empty on both sides", both_empty),
        ("Row counts MISMATCH", mismatched),
        ("Errors", errors),
        ("Log file", LOG_FILE),
        ("Generated at", time.strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(rows, 1):
        s.cell(row=i, column=1, value=k).font = Font(bold=True)
        s.cell(row=i, column=2, value=v)
    s.column_dimensions["A"].width = 34
    s.column_dimensions["B"].width = 40

    wb.save(path)
    log.info("Report written: %s", path)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    t0 = time.time()
    log.info("=" * 70)
    log.info("Starting comparison for %d schema(s), %d workers per DB",
             len(SCHEMAS), MAX_WORKERS_PER_DB)
    log.info("Log file: %s (DEBUG level, full SQL + responses)", LOG_FILE)
    log.info("=" * 70)
    results = run_comparison()
    write_report(results, OUTPUT_FILE)

    mismatches = [r for r in results if r.match != "YES"]
    log.info("Done in %.1fs. %d/%d tables fully match.",
             time.time() - t0, len(results) - len(mismatches), len(results))
    if mismatches:
        log.warning("%d table(s) need attention -- see %s and %s",
                    len(mismatches), OUTPUT_FILE, LOG_FILE)
        sys.exit(1)


if __name__ == "__main__":
    main()
