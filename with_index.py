#!/usr/bin/env python3
"""
Multi-threaded Oracle vs PostgreSQL (Aurora) comparison tool.

Compares, per schema pair:
  * Row counts for every table (union of both sides).
  * Indexes on tables that exist on both sides -- matched by column signature
    (ordered list of key columns), not by name.

Output: an XLSX report with three sheets:
    Row Count Comparison | Index Comparison | Summary

Requirements:
    pip install oracledb psycopg2-binary openpyxl
"""

import concurrent.futures as cf
import logging
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Optional, Tuple

import oracledb
import psycopg2
from psycopg2 import pool as pg_pool
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# CONFIG
# ----------------------------------------------------------------------------
ORACLE_CONFIG = {
    "user":     "oracle_user",
    "password": "oracle_password",
    "dsn":      "oracle-host.example.com:1521/ORCLPDB1",
}

POSTGRES_CONFIG = {
    "user":     "pg_user",
    "password": "pg_password",
    "host":     "aurora-cluster.cluster-xxxx.us-east-1.rds.amazonaws.com",
    "port":     5432,
    "dbname":   "mydb",
}

SCHEMA_PAIRS = [
    ("SALES",   "sales"),
    ("HR",      "hr"),
    ("FINANCE", "finance"),
]

MAX_WORKERS_PER_DB = 16
POOL_SIZE = MAX_WORKERS_PER_DB
COMPARE_INDEXES = True
OUTPUT_FILE = "db_comparison_report.xlsx"

# ----------------------------------------------------------------------------
# Logging
# ----------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(threadName)-12s] %(levelname)-7s %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("compare")

# ----------------------------------------------------------------------------
# Data models
# ----------------------------------------------------------------------------
@dataclass
class TableResult:
    oracle_schema: str
    pg_schema: str
    table_name: str
    in_oracle: bool = False
    in_postgres: bool = False
    ora_table_actual: str = ""
    pg_table_actual: str = ""
    oracle_count: Optional[int] = None
    pg_count: Optional[int] = None
    oracle_error: str = ""
    pg_error: str = ""

    @property
    def exists_both(self) -> str:
        return "YES" if (self.in_oracle and self.in_postgres) else "NO"

    @property
    def match(self) -> str:
        if not (self.in_oracle and self.in_postgres):
            return "NO"
        if self.oracle_count is None or self.pg_count is None:
            return "ERROR"
        return "YES" if self.oracle_count == self.pg_count else "NO"


@dataclass
class IndexInfo:
    """One index (either Oracle or Postgres)."""
    schema: str
    table_key: str                   # table name, lowercased, for matching
    index_name: str
    unique: bool
    columns: Tuple[str, ...]         # lowercased, ordered

    @property
    def signature(self) -> Tuple[str, ...]:
        return self.columns


@dataclass
class IndexRow:
    """One row in the index-comparison report."""
    oracle_schema: str
    pg_schema: str
    table_name: str
    columns: str
    oracle_index: str
    pg_index: str
    oracle_unique: str      # "YES" / "NO" / "-"
    pg_unique: str
    match: str              # "YES" / "NO"
    remarks: str


# ----------------------------------------------------------------------------
# Connection pools
# ----------------------------------------------------------------------------
class OraclePool:
    def __init__(self, cfg, size):
        self.pool = oracledb.create_pool(
            user=cfg["user"], password=cfg["password"], dsn=cfg["dsn"],
            min=2, max=size, increment=1,
            getmode=oracledb.POOL_GETMODE_WAIT,
        )

    def query_one(self, sql, params=None):
        conn = self.pool.acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchone()
        finally:
            self.pool.release(conn)

    def query_all(self, sql, params=None):
        conn = self.pool.acquire()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params or {})
                return cur.fetchall()
        finally:
            self.pool.release(conn)

    def close(self):
        self.pool.close(force=True)


class PostgresPool:
    def __init__(self, cfg, size):
        self.pool = pg_pool.ThreadedConnectionPool(minconn=2, maxconn=size, **cfg)

    def _run(self, sql, params, fetch):
        conn = self.pool.getconn()
        try:
            with conn.cursor() as cur:
                cur.execute(sql, params)
                return fetch(cur)
        finally:
            conn.rollback()
            self.pool.putconn(conn)

    def query_one(self, sql, params=None):
        return self._run(sql, params, lambda c: c.fetchone())

    def query_all(self, sql, params=None):
        return self._run(sql, params, lambda c: c.fetchall())

    def close(self):
        self.pool.closeall()


# ----------------------------------------------------------------------------
# Table metadata
# ----------------------------------------------------------------------------
def get_oracle_tables(orapool: OraclePool, schema: str) -> set:
    rows = orapool.query_all(
        """
        SELECT table_name
        FROM all_tables
        WHERE owner = :owner
          AND table_name NOT LIKE 'BIN$%'
          AND (nested = 'NO' OR nested IS NULL)
          AND iot_type IS NULL
        """,
        {"owner": schema.upper()},
    )
    tables = {r[0] for r in rows}
    log.info("Oracle schema %s: %d tables found", schema, len(tables))
    return tables


def get_postgres_tables(pgpool: PostgresPool, schema: str) -> set:
    rows = pgpool.query_all(
        """
        SELECT c.relname
        FROM pg_class c
        JOIN pg_namespace n ON n.oid = c.relnamespace
        WHERE n.nspname = %s
          AND c.relkind IN ('r', 'p', 'f')
          AND c.relname NOT LIKE 'pg_%%'
        """,
        (schema,),
    )
    tables = {r[0] for r in rows}
    log.info("Postgres schema %s: %d tables found", schema, len(tables))
    return tables


# ----------------------------------------------------------------------------
# Index metadata
# ----------------------------------------------------------------------------
def get_oracle_indexes(orapool: OraclePool, schema: str) -> List[IndexInfo]:
    rows = orapool.query_all(
        """
        SELECT i.table_name,
               i.index_name,
               i.uniqueness,
               c.column_name,
               c.column_position
        FROM all_indexes i
        JOIN all_ind_columns c
          ON i.owner = c.index_owner
         AND i.index_name = c.index_name
        WHERE i.owner = :owner
          AND i.index_name NOT LIKE 'BIN$%'
          AND i.dropped = 'NO'
          AND i.index_type NOT LIKE 'IOT%'
          AND i.index_type NOT LIKE 'LOB%'
          AND i.table_name NOT LIKE 'BIN$%'
        ORDER BY i.table_name, i.index_name, c.column_position
        """,
        {"owner": schema.upper()},
    )
    # Aggregate rows into IndexInfo per (table, index).
    bucket = {}
    for tbl, idx, uniq, col, _pos in rows:
        key = (tbl, idx)
        info = bucket.get(key)
        if info is None:
            info = IndexInfo(
                schema=schema,
                table_key=tbl.lower(),
                index_name=idx,
                unique=(uniq == "UNIQUE"),
                columns=[],
            )
            bucket[key] = info
        info.columns.append((col or "").lower())
    result = []
    for info in bucket.values():
        info.columns = tuple(info.columns)
        result.append(info)
    log.info("Oracle schema %s: %d indexes found", schema, len(result))
    return result


def get_postgres_indexes(pgpool: PostgresPool, schema: str) -> List[IndexInfo]:
    # Only key columns are used for the signature. INCLUDE / covering
    # columns (present when indnkeyatts < indnatts, PG 11+) are excluded.
    rows = pgpool.query_all(
        """
        SELECT tc.relname          AS table_name,
               ic.relname          AS index_name,
               ix.indisunique      AS is_unique,
               col.attname         AS column_name,
               col.ord             AS column_position
        FROM pg_index ix
        JOIN pg_class ic ON ic.oid = ix.indexrelid
        JOIN pg_class tc ON tc.oid = ix.indrelid
        JOIN pg_namespace n ON n.oid = tc.relnamespace
        CROSS JOIN LATERAL (
            SELECT COALESCE(a.attname, '<expr>') AS attname, k.ord
            FROM unnest(ix.indkey) WITH ORDINALITY AS k(attnum, ord)
            LEFT JOIN pg_attribute a
                   ON a.attrelid = tc.oid
                  AND a.attnum   = k.attnum
            WHERE k.ord <= ix.indnkeyatts
        ) col
        WHERE n.nspname = %s
          AND tc.relkind IN ('r', 'p')
          AND tc.relname NOT LIKE 'pg_%%'
        ORDER BY tc.relname, ic.relname, col.ord
        """,
        (schema,),
    )
    bucket = {}
    for tbl, idx, uniq, col, _pos in rows:
        key = (tbl, idx)
        info = bucket.get(key)
        if info is None:
            info = IndexInfo(
                schema=schema,
                table_key=tbl.lower(),
                index_name=idx,
                unique=bool(uniq),
                columns=[],
            )
            bucket[key] = info
        info.columns.append((col or "").lower())
    result = []
    for info in bucket.values():
        info.columns = tuple(info.columns)
        result.append(info)
    log.info("Postgres schema %s: %d indexes found", schema, len(result))
    return result


# ----------------------------------------------------------------------------
# Row-count workers
# ----------------------------------------------------------------------------
PG_UNDEFINED_TABLE = "42P01"
ORA_TABLE_OR_VIEW_NOT_EXIST = 942


def count_oracle(orapool: OraclePool, result: TableResult):
    sql = f'SELECT COUNT(*) FROM "{result.oracle_schema.upper()}"."{result.ora_table_actual}"'
    t0 = time.time()
    try:
        row = orapool.query_one(sql)
        result.oracle_count = int(row[0])
        log.info("Oracle  %s.%s = %s (%.1fs)",
                 result.oracle_schema, result.ora_table_actual,
                 f"{result.oracle_count:,}", time.time() - t0)
    except oracledb.DatabaseError as e:
        err, = e.args
        if getattr(err, "code", None) == ORA_TABLE_OR_VIEW_NOT_EXIST:
            result.in_oracle = False
            result.oracle_error = "Table not found at count time"
        else:
            result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle  %s.%s FAILED: %s",
                  result.oracle_schema, result.ora_table_actual,
                  result.oracle_error)
    except Exception as e:
        result.oracle_error = str(e).splitlines()[0][:200]
        log.error("Oracle  %s.%s FAILED: %s",
                  result.oracle_schema, result.ora_table_actual,
                  result.oracle_error)


def count_postgres(pgpool: PostgresPool, result: TableResult):
    sql = f'SELECT COUNT(*) FROM "{result.pg_schema}"."{result.pg_table_actual}"'
    t0 = time.time()
    try:
        row = pgpool.query_one(sql)
        result.pg_count = int(row[0])
        log.info("Postgres %s.%s = %s (%.1fs)",
                 result.pg_schema, result.pg_table_actual,
                 f"{result.pg_count:,}", time.time() - t0)
    except psycopg2.Error as e:
        if getattr(e, "pgcode", None) == PG_UNDEFINED_TABLE:
            result.in_postgres = False
            result.pg_error = "Table not found at count time"
        else:
            result.pg_error = str(e).splitlines()[0][:200]
        log.error("Postgres %s.%s FAILED: %s",
                  result.pg_schema, result.pg_table_actual,
                  result.pg_error)
    except Exception as e:
        result.pg_error = str(e).splitlines()[0][:200]
        log.error("Postgres %s.%s FAILED: %s",
                  result.pg_schema, result.pg_table_actual,
                  result.pg_error)


# ----------------------------------------------------------------------------
# Index comparison logic
# ----------------------------------------------------------------------------
def compare_indexes_for_pair(ora_schema: str, pg_schema: str,
                             common_tables: set,
                             ora_indexes: List[IndexInfo],
                             pg_indexes: List[IndexInfo]) -> List[IndexRow]:
    """Compare indexes for tables that exist on both sides."""
    # Group indexes by table.
    ora_by_table = defaultdict(list)
    for idx in ora_indexes:
        if idx.table_key in common_tables:
            ora_by_table[idx.table_key].append(idx)
    pg_by_table = defaultdict(list)
    for idx in pg_indexes:
        if idx.table_key in common_tables:
            pg_by_table[idx.table_key].append(idx)

    rows: List[IndexRow] = []
    for table_key in sorted(common_tables):
        ora_list = ora_by_table.get(table_key, [])
        pg_list = pg_by_table.get(table_key, [])
        if not ora_list and not pg_list:
            continue

        # Group each side by column signature. Multiple indexes with the
        # same signature (rare) are paired in list order.
        ora_by_sig = defaultdict(list)
        for i in ora_list:
            ora_by_sig[i.signature].append(i)
        pg_by_sig = defaultdict(list)
        for i in pg_list:
            pg_by_sig[i.signature].append(i)

        display_table = ora_list[0].table_key if ora_list else pg_list[0].table_key
        # Prefer the original-case name if we have one.
        if ora_list:
            # Rebuild table name from any index's stored schema info -- we
            # only kept the lower-cased key, so use the key here; the row
            # count sheet has the exact casing.
            display_table = ora_list[0].table_key
        else:
            display_table = pg_list[0].table_key

        for sig in sorted(set(ora_by_sig) | set(pg_by_sig)):
            o = ora_by_sig.get(sig, [])
            p = pg_by_sig.get(sig, [])
            pairs = max(len(o), len(p))
            for k in range(pairs):
                oi = o[k] if k < len(o) else None
                pi = p[k] if k < len(p) else None
                columns_display = ", ".join(sig) if sig else "(none)"

                remarks = []
                if oi and not pi:
                    match = "NO"
                    remarks.append("Missing in Postgres")
                elif pi and not oi:
                    match = "NO"
                    remarks.append("Missing in Oracle")
                else:
                    # Both exist -- compare uniqueness.
                    if oi.unique == pi.unique:
                        match = "YES"
                    else:
                        match = "NO"
                        remarks.append("Uniqueness differs")
                if sig and any(c == "<expr>" or c.startswith("sys_nc")
                               for c in sig):
                    remarks.append("Contains expression/function column")

                rows.append(IndexRow(
                    oracle_schema=ora_schema,
                    pg_schema=pg_schema,
                    table_name=display_table,
                    columns=columns_display,
                    oracle_index=oi.index_name if oi else "-",
                    pg_index=pi.index_name if pi else "-",
                    oracle_unique=("YES" if oi.unique else "NO") if oi else "-",
                    pg_unique=("YES" if pi.unique else "NO") if pi else "-",
                    match=match,
                    remarks="; ".join(remarks),
                ))
    return rows


# ----------------------------------------------------------------------------
# Orchestration
# ----------------------------------------------------------------------------
@dataclass
class SchemaBundle:
    """All metadata + results for one schema pair."""
    ora_schema: str
    pg_schema: str
    ora_tables: set = field(default_factory=set)
    pg_tables: set = field(default_factory=set)
    ora_indexes: List[IndexInfo] = field(default_factory=list)
    pg_indexes: List[IndexInfo] = field(default_factory=list)
    table_results: List[TableResult] = field(default_factory=list)
    index_results: List[IndexRow] = field(default_factory=list)


def gather_metadata(orapool, pgpool, meta_ex) -> List[SchemaBundle]:
    """Fetch tables + indexes for every schema pair in parallel."""
    bundles = [SchemaBundle(o, p) for o, p in SCHEMA_PAIRS]
    futs = {}
    for b in bundles:
        futs[meta_ex.submit(get_oracle_tables, orapool, b.ora_schema)] = (b, "ora_tables")
        futs[meta_ex.submit(get_postgres_tables, pgpool, b.pg_schema)] = (b, "pg_tables")
        if COMPARE_INDEXES:
            futs[meta_ex.submit(get_oracle_indexes, orapool, b.ora_schema)] = (b, "ora_indexes")
            futs[meta_ex.submit(get_postgres_indexes, pgpool, b.pg_schema)] = (b, "pg_indexes")

    for fut in cf.as_completed(futs):
        bundle, attr = futs[fut]
        setattr(bundle, attr, fut.result())

    # Build TableResult objects from tables sets.
    for b in bundles:
        ora_map = {t.lower(): t for t in b.ora_tables}
        pg_map = {t.lower(): t for t in b.pg_tables}
        for key in sorted(set(ora_map) | set(pg_map)):
            b.table_results.append(TableResult(
                oracle_schema=b.ora_schema,
                pg_schema=b.pg_schema,
                table_name=ora_map.get(key, pg_map.get(key, key)),
                in_oracle=key in ora_map,
                in_postgres=key in pg_map,
                ora_table_actual=ora_map.get(key, ""),
                pg_table_actual=pg_map.get(key, ""),
            ))
    return bundles


def run_comparison():
    orapool = OraclePool(ORACLE_CONFIG, POOL_SIZE)
    pgpool = PostgresPool(POSTGRES_CONFIG, POOL_SIZE)
    bundles: List[SchemaBundle] = []

    try:
        # Phase 1: all metadata in parallel.
        meta_workers = max(len(SCHEMA_PAIRS) * 4, 4)
        with cf.ThreadPoolExecutor(max_workers=meta_workers,
                                   thread_name_prefix="meta") as meta_ex:
            bundles = gather_metadata(orapool, pgpool, meta_ex)

        all_table_results = [r for b in bundles for r in b.table_results]
        log.info("Total tables (union): %d", len(all_table_results))

        # Phase 2: row counts on two dedicated pools.
        with cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="ora") as ora_ex, \
             cf.ThreadPoolExecutor(max_workers=MAX_WORKERS_PER_DB,
                                   thread_name_prefix="pg") as pg_ex:
            futures = []
            for r in all_table_results:
                if r.in_oracle:
                    futures.append(ora_ex.submit(count_oracle, orapool, r))
                if r.in_postgres:
                    futures.append(pg_ex.submit(count_postgres, pgpool, r))
            done, total = 0, len(futures)
            for fut in cf.as_completed(futures):
                fut.result()
                done += 1
                if done % 25 == 0 or done == total:
                    log.info("Progress: %d/%d count queries complete",
                             done, total)

        # Phase 3: index comparison (CPU-only, cheap).
        if COMPARE_INDEXES:
            for b in bundles:
                common = ({t.lower() for t in b.ora_tables}
                          & {t.lower() for t in b.pg_tables})
                b.index_results = compare_indexes_for_pair(
                    b.ora_schema, b.pg_schema, common,
                    b.ora_indexes, b.pg_indexes,
                )
            total_idx = sum(len(b.index_results) for b in bundles)
            log.info("Index comparison rows: %d", total_idx)
    finally:
        orapool.close()
        pgpool.close()

    all_table_results = [r for b in bundles for r in b.table_results]
    all_table_results.sort(key=lambda r: (r.oracle_schema, r.table_name.lower()))
    all_index_results = [r for b in bundles for r in b.index_results]
    all_index_results.sort(key=lambda r: (r.oracle_schema, r.table_name, r.columns))
    return all_table_results, all_index_results


# ----------------------------------------------------------------------------
# XLSX report
# ----------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
THIN = Border(*[Side(style="thin", color="D0D0D0")] * 4)
CENTER = Alignment(horizontal="center")


def _apply_headers(ws, headers, widths):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, CENTER, THIN
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"


def write_row_count_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Exists",
               "Oracle_Count", "Postgres_Count", "Match", "Remarks"]
    _apply_headers(ws, headers, [18, 18, 38, 10, 16, 16, 10, 55])

    for i, r in enumerate(results, start=2):
        remarks = []
        if not r.in_oracle and r.in_postgres:
            remarks.append("Missing in Oracle")
        elif r.in_oracle and not r.in_postgres:
            remarks.append("Missing in Postgres")
        elif r.in_oracle and r.in_postgres \
                and r.oracle_count == 0 and r.pg_count == 0:
            remarks.append("Both sides empty (0 rows)")
        if r.oracle_error:
            remarks.append(f"Oracle: {r.oracle_error}")
        if r.pg_error:
            remarks.append(f"Postgres: {r.pg_error}")

        ora_val = r.oracle_count if r.in_oracle and r.oracle_count is not None else "N/A"
        pg_val = r.pg_count if r.in_postgres and r.pg_count is not None else "N/A"

        row_vals = [r.oracle_schema, r.pg_schema, r.table_name,
                    r.exists_both, ora_val, pg_val, r.match,
                    "; ".join(remarks)]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN
            if col in (4, 7):
                c.alignment = CENTER
            if col in (5, 6) and isinstance(v, int) and not isinstance(v, bool):
                c.number_format = "#,##0"

        match_cell = ws.cell(row=i, column=7)
        exists_cell = ws.cell(row=i, column=4)
        if r.match == "YES":
            match_cell.fill = GREEN
        elif r.match == "ERROR":
            match_cell.fill = YELLOW
        else:
            match_cell.fill = RED
        exists_cell.fill = GREEN if r.exists_both == "YES" else RED

    if results:
        ws.auto_filter.ref = f"A1:H{len(results) + 1}"


def write_index_sheet(ws, results):
    headers = ["Oracle_Schema", "Postgres_Schema", "Table_Name", "Columns",
               "Oracle_Index", "Postgres_Index",
               "Oracle_Unique", "Postgres_Unique", "Match", "Remarks"]
    _apply_headers(ws, headers, [16, 16, 30, 40, 30, 30, 10, 10, 10, 40])

    for i, r in enumerate(results, start=2):
        row_vals = [r.oracle_schema, r.pg_schema, r.table_name, r.columns,
                    r.oracle_index, r.pg_index,
                    r.oracle_unique, r.pg_unique, r.match, r.remarks]
        for col, v in enumerate(row_vals, 1):
            c = ws.cell(row=i, column=col, value=v)
            c.border = THIN
            if col in (7, 8, 9):
                c.alignment = CENTER
        match_cell = ws.cell(row=i, column=9)
        match_cell.fill = GREEN if r.match == "YES" else RED

    if results:
        ws.auto_filter.ref = f"A1:J{len(results) + 1}"


def write_summary_sheet(ws, table_results, index_results):
    def count(cond, seq):
        return sum(1 for r in seq if cond(r))

    total = len(table_results)
    both = count(lambda r: r.exists_both == "YES", table_results)
    matched = count(lambda r: r.match == "YES", table_results)
    mismatched = count(lambda r: r.match == "NO" and r.exists_both == "YES",
                       table_results)
    missing_pg = count(lambda r: r.in_oracle and not r.in_postgres,
                       table_results)
    missing_ora = count(lambda r: not r.in_oracle and r.in_postgres,
                        table_results)
    errors = count(lambda r: r.match == "ERROR", table_results)
    both_empty = count(
        lambda r: (r.in_oracle and r.in_postgres
                   and r.oracle_count == 0 and r.pg_count == 0),
        table_results,
    )

    idx_total = len(index_results)
    idx_matched = count(lambda r: r.match == "YES", index_results)
    idx_missing_pg = count(lambda r: "Missing in Postgres" in r.remarks,
                           index_results)
    idx_missing_ora = count(lambda r: "Missing in Oracle" in r.remarks,
                            index_results)
    idx_uniq_diff = count(lambda r: "Uniqueness differs" in r.remarks,
                          index_results)

    rows = [
        ("=== Tables ===", ""),
        ("Total tables (union)", total),
        ("Exists in both", both),
        ("Missing in Postgres only", missing_pg),
        ("Missing in Oracle only", missing_ora),
        ("Row counts MATCH", matched),
        ("  ...of which empty on both sides", both_empty),
        ("Row counts MISMATCH", mismatched),
        ("Errors", errors),
        ("", ""),
        ("=== Indexes ===", ""),
        ("Total index comparisons", idx_total),
        ("Indexes matched", idx_matched),
        ("Missing in Postgres", idx_missing_pg),
        ("Missing in Oracle", idx_missing_ora),
        ("Uniqueness differs", idx_uniq_diff),
        ("", ""),
        ("Generated at", time.strftime("%Y-%m-%d %H:%M:%S")),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22


def write_report(table_results, index_results, path):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Row Count Comparison"
    write_row_count_sheet(ws1, table_results)

    if COMPARE_INDEXES:
        ws2 = wb.create_sheet("Index Comparison")
        write_index_sheet(ws2, index_results)

    ws3 = wb.create_sheet("Summary")
    write_summary_sheet(ws3, table_results, index_results)

    wb.save(path)
    log.info("Report written: %s", path)


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    t0 = time.time()
    log.info("Starting comparison for %d schema pair(s), %d workers per DB",
             len(SCHEMA_PAIRS), MAX_WORKERS_PER_DB)
    table_results, index_results = run_comparison()
    write_report(table_results, index_results, OUTPUT_FILE)

    tbl_bad = [r for r in table_results if r.match != "YES"]
    idx_bad = [r for r in index_results if r.match != "YES"]
    log.info("Done in %.1fs. Tables: %d/%d match. Indexes: %d/%d match.",
             time.time() - t0,
             len(table_results) - len(tbl_bad), len(table_results),
             len(index_results) - len(idx_bad), len(index_results))
    if tbl_bad or idx_bad:
        log.warning("Issues found -- see %s", OUTPUT_FILE)
        sys.exit(1)


if __name__ == "__main__":
    main()
